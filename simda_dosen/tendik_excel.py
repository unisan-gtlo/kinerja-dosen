"""Ekspor Excel Daftar Data Tendik (Kelola Data Tendik)."""
import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def render_excel_daftar_tendik(daftar, *, kata_kunci=""):
    """Return io.BytesIO berisi file .xlsx. `daftar` queryset/list
    DataTendik -- fungsi ini murni tampilan, tidak melakukan query apa
    pun."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Tendik"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "DAFTAR DATA TENDIK"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = "Universitas Ichsan Gorontalo"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = center

    row_header = 4
    if kata_kunci:
        ws.merge_cells("A3:H3")
        ws["A3"] = f'Kata kunci pencarian: "{kata_kunci}"'
        ws["A3"].font = Font(size=9, italic=True)
        ws["A3"].alignment = center
        row_header = 5

    headers = ["No", "Nama", "NITK", "NIP Yayasan", "Unit Kerja", "Status Kepegawaian", "Pendidikan Terakhir", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 26

    for idx, t in enumerate(daftar, 1):
        row_data = [
            idx, t.nama_lengkap, t.nip or "-", t.nip_yayasan or "-",
            t.unit_kerja_nama or "-", t.status_kepegawaian_nama or "-",
            t.get_pendidikan_terakhir_display() if t.pendidikan_terakhir else "-",
            "Aktif" if t.is_active else "Nonaktif",
        ]
        row_num = row_header + idx
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            cell.alignment = center if col != 2 else left

    col_widths = [5, 32, 16, 18, 28, 24, 20, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
