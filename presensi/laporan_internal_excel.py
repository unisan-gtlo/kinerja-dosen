"""Ekspor Excel Laporan Presensi Internal -- versi openpyxl dari
presensi/laporan_internal_pdf.py, layout & data sama persis."""
import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .laporan_internal import data_laporan_internal
from .laporan_internal_pdf import _label_filter
from .laporan_serdos import jenis_tanggal_bulan
from .models import NAMA_BULAN

WARNA_HEADER = "1e3a5f"
WARNA_ARSIR = "D9D9D9"


def render_excel_laporan_internal(user_qs, bulan, tahun, kategori="", fakultas="", prodi=""):
    """Return io.BytesIO berisi file .xlsx."""
    daftar = data_laporan_internal(user_qs, bulan, tahun)
    jenis_list = jenis_tanggal_bulan(bulan, tahun)
    jumlah_hari = len(jenis_list)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Presensi"

    header_font = Font(bold=True, color="FFFFFF", size=9)
    header_fill = PatternFill(start_color=WARNA_HEADER, end_color=WARNA_HEADER, fill_type="solid")
    arsir_fill = PatternFill(start_color=WARNA_ARSIR, end_color=WARNA_ARSIR, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),
    )

    jumlah_kolom = 3 + jumlah_hari  # No, Nama, Program Studi, tanggal...
    kolom_terakhir = get_column_letter(jumlah_kolom)

    ws.merge_cells(f"A1:{kolom_terakhir}1")
    ws["A1"] = "LAPORAN PRESENSI BULANAN"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    ws.merge_cells(f"A2:{kolom_terakhir}2")
    ws["A2"] = f"Universitas Ichsan Gorontalo · {NAMA_BULAN[bulan - 1]} {tahun}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = center

    label_filter = _label_filter(kategori, fakultas, prodi)
    baris_berikut = 3
    if label_filter:
        ws.merge_cells(f"A3:{kolom_terakhir}3")
        ws["A3"] = label_filter
        ws["A3"].font = Font(size=9, italic=True)
        ws["A3"].alignment = center
        baris_berikut = 4

    row_header = baris_berikut + 1
    headers = ["No", "Nama", "Program Studi"] + [str(tanggal.day) for tanggal, _ in jenis_list]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 20

    for idx, baris in enumerate(daftar, 1):
        row_num = row_header + idx
        ws.row_dimensions[row_num].height = 24

        ws.cell(row=row_num, column=1, value=idx).alignment = center
        nama_cell = ws.cell(row=row_num, column=2, value=baris["user"].get_full_name() or baris["user"].username)
        nama_cell.alignment = left
        ws.cell(row=row_num, column=3, value=baris["nama_prodi"]).alignment = center

        for i, hari in enumerate(baris["hari_grid"]):
            col = 4 + i
            teks = f"{hari.jam_masuk}\n{hari.jam_pulang}" if (hari.jam_masuk or hari.jam_pulang) else ""
            ws.cell(row=row_num, column=col, value=teks).alignment = center

        for col in range(1, jumlah_kolom + 1):
            ws.cell(row=row_num, column=col).border = thin

    baris_terakhir = row_header + len(daftar)
    for i, (_tanggal, jenis) in enumerate(jenis_list):
        if jenis != "kerja" and daftar:
            col = 4 + i
            for row_num in range(row_header + 1, baris_terakhir + 1):
                ws.cell(row=row_num, column=col).fill = arsir_fill

    col_widths = [5, 24, 14] + [5] * jumlah_hari
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    baris_ket = baris_terakhir + 2
    ws.cell(row=baris_ket, column=1).fill = arsir_fill
    ws.cell(
        row=baris_ket, column=2,
        value=": Hari Minggu/Libur -- jam kosong berarti tidak ada presensi tercatat.",
    ).font = Font(size=9)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
