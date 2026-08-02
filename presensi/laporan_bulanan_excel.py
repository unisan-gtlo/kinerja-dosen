"""Ekspor Excel Laporan Bulanan Jam Kerja -- dipindah dari views.py supaya
konsisten dengan pola modul render_* terpisah (laporan_internal_excel.py,
laporan_detail_excel.py). Ditambah label filter kategori/fakultas/prodi
dan blok pengesahan opsional sejak Laporan Bulanan digabung satu halaman
dengan Laporan Presensi Internal (2026-08-02) -- keduanya sekarang pakai
form & blok tanda tangan yang sama."""
import io

import openpyxl
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .laporan_internal_pdf import _label_filter
from .models import NAMA_BULAN


def render_excel_laporan_bulanan(
    daftar, bulan, tahun, *, kategori="", fakultas="", prodi="",
    kota="Gorontalo", tanggal_cetak=None, jabatan_penandatangan="", nama_penandatangan="", id_penandatangan="",
):
    """Return io.BytesIO berisi file .xlsx. `daftar` sudah hasil
    rekap.laporan_bulanan_jam_kerja() -- fungsi ini murni tampilan."""
    tanggal_cetak = tanggal_cetak or timezone.localdate()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "LAPORAN BULANAN JAM KERJA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Universitas Ichsan Gorontalo · {NAMA_BULAN[bulan - 1]} {tahun}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = center

    row_header = 4
    label_filter = _label_filter(kategori, fakultas, prodi)
    if label_filter:
        ws.merge_cells("A3:H3")
        ws["A3"] = label_filter
        ws["A3"].font = Font(size=9, italic=True)
        ws["A3"].alignment = center
        row_header = 5

    headers = ["No", "Nama", "Role", "Kelompok", "Hari Hadir", "Total Jam Kerja", "Target Jam Kerja", "Selisih"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 26

    for idx, item in enumerate(daftar, 1):
        u = item["user"]
        target = item["target"]
        row_data = [
            idx,
            u.nama_resmi,
            u.get_role_display_id(),
            item["kelompok"].nama if item["kelompok"] else "-",
            item["hari_hadir"],
            item["total_jam_kerja"],
            f"{target.target_jam_kerja} jam" if target else "-",
            item["selisih_jam_kerja"] or "-",
        ]
        row_num = row_header + idx
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            cell.alignment = center if col != 2 else left

    col_widths = [5, 28, 16, 16, 10, 14, 14, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    if jabatan_penandatangan or nama_penandatangan or id_penandatangan:
        baris_ttd = row_header + len(daftar) + 3
        ws.cell(row=baris_ttd, column=6, value=f"{kota}, {tanggal_cetak.strftime('%d %B %Y').upper()}")
        ws.cell(row=baris_ttd, column=6).alignment = center
        if jabatan_penandatangan:
            ws.cell(row=baris_ttd + 1, column=6, value=f"{jabatan_penandatangan.upper()},")
            ws.cell(row=baris_ttd + 1, column=6).alignment = center

        baris_nama_ttd = baris_ttd + 5
        nama_cell = ws.cell(row=baris_nama_ttd, column=6, value=nama_penandatangan or ".........................")
        nama_cell.font = Font(underline="single")
        nama_cell.alignment = center
        if id_penandatangan:
            id_cell = ws.cell(row=baris_nama_ttd + 1, column=6, value=id_penandatangan)
            id_cell.alignment = center

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
