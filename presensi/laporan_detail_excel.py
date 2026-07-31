"""Ekspor Excel Detail Presensi Bulanan -- versi openpyxl dari
presensi/laporan_detail_pdf.py."""
import io

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .laporan_detail import detail_presensi_bulanan
from .models import NAMA_BULAN

WARNA_HEADER = "1e3a5f"
WARNA_ARSIR = "D9D9D9"


def render_excel_detail_presensi(user, bulan, tahun):
    """Return io.BytesIO berisi file .xlsx."""
    data = detail_presensi_bulanan(user, bulan, tahun)
    rekap = data["rekap"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Detail Presensi"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=WARNA_HEADER, end_color=WARNA_HEADER, fill_type="solid")
    arsir_fill = PatternFill(start_color=WARNA_ARSIR, end_color=WARNA_ARSIR, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:G1")
    ws["A1"] = "DETAIL PRESENSI BULANAN"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    nama_tampil = user.get_full_name() or user.username
    if user.nidn:
        nama_tampil += f" ({user.nidn})"
    ws.merge_cells("A2:G2")
    ws["A2"] = f"{nama_tampil} · {NAMA_BULAN[bulan - 1]} {tahun}"
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = center

    row_header = 4
    headers = ["Tanggal", "Hari", "Status", "Masuk", "Pulang", "Keterangan", "Durasi"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 20

    for idx, baris in enumerate(data["baris"], 1):
        row_num = row_header + idx
        row_data = [
            baris.tanggal.strftime("%d-%m-%Y"), baris.nama_hari, baris.status,
            baris.waktu_masuk, baris.waktu_pulang, baris.keterangan, baris.durasi_kerja,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            cell.alignment = left if col == 6 else center
            if baris.jenis != "kerja":
                cell.fill = arsir_fill

    col_widths = [12, 10, 12, 9, 9, 40, 9]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    baris_ringkasan = row_header + len(data["baris"]) + 2
    ws.cell(row=baris_ringkasan, column=1, value="Hari Hadir").font = Font(bold=True)
    ws.cell(row=baris_ringkasan, column=2, value=f"{rekap['hari_hadir']} hari")
    ws.cell(row=baris_ringkasan + 1, column=1, value="Total Jam Kerja").font = Font(bold=True)
    ws.cell(row=baris_ringkasan + 1, column=2, value=rekap["total_jam_kerja"])
    if rekap["target"]:
        ws.cell(row=baris_ringkasan + 2, column=1, value=f"Target ({rekap['kelompok'].nama})").font = Font(bold=True)
        ws.cell(
            row=baris_ringkasan + 2, column=2,
            value=f"{rekap['target'].target_jam_kerja} jam / {rekap['target'].target_hari_kerja} hari",
        )
        ws.cell(row=baris_ringkasan + 3, column=1, value="Selisih").font = Font(bold=True)
        ws.cell(row=baris_ringkasan + 3, column=2, value=rekap["selisih_jam_kerja"])

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
