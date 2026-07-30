"""Ekspor Excel Laporan Daftar Hadir Dosen Serdos (LLDIKTI) -- versi
openpyxl dari presensi/laporan_serdos_pdf.py, layout & data sama persis,
cuma format file beda (institusi kadang butuh versi Excel yang bisa
diedit/direkap ulang, bukan cuma PDF final)."""
import io

import openpyxl
from django.contrib.staticfiles import finders
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .laporan_serdos import data_laporan_serdos, jenis_tanggal_bulan
from .laporan_serdos_pdf import NAMA_LLDIKTI, NAMA_PTS, NAMA_YAYASAN
from .models import NAMA_BULAN, HariLibur

WARNA_HEADER = "1e3a5f"
WARNA_ARSIR = "D9D9D9"

LEBAR_PARAF_PX = 45
TINGGI_PARAF_PX = 24


def _gambar_aman(field_file):
    """openpyxl Image dari FieldFile Django, atau None kalau kosong/rusak
    -- sama seperti presensi/laporan_serdos_pdf.py::_gambar_aman."""
    if not field_file:
        return None
    try:
        img = XLImage(field_file.path)
        img.width = LEBAR_PARAF_PX
        img.height = TINGGI_PARAF_PX
        return img
    except (FileNotFoundError, OSError):
        return None


def render_excel_laporan_serdos(
    bulan, tahun, kota, tanggal_cetak,
    jabatan_penandatangan="Rektor", nama_penandatangan="", nip_penandatangan="",
    file_ttd_penandatangan=None,
):
    """Return openpyxl.Workbook -- pemanggil (presensi/views.py) yang
    urus HttpResponse & Content-Disposition."""
    daftar = data_laporan_serdos(bulan, tahun)
    jenis_list = jenis_tanggal_bulan(bulan, tahun)
    jumlah_hari = len(jenis_list)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Daftar Hadir Serdos"

    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color=WARNA_HEADER, end_color=WARNA_HEADER, fill_type="solid")
    arsir_fill = PatternFill(start_color=WARNA_ARSIR, end_color=WARNA_ARSIR, fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"),
    )

    jumlah_kolom = 3 + jumlah_hari + 1  # No, Nama/NIDN, Gol/Jabatan, tanggal..., Total Hadir
    kolom_terakhir = get_column_letter(jumlah_kolom)

    logo_path = finders.find("img/kampus.png")
    if logo_path:
        logo = XLImage(logo_path)
        logo.width, logo.height = 60, 60
        ws.add_image(logo, "A1")

    ws.merge_cells(f"A1:{kolom_terakhir}1")
    ws["A1"] = "DAFTAR HADIR DOSEN TETAP YAYASAN PENERIMA SERDOS"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A1"].alignment = center

    ws.merge_cells(f"A2:{kolom_terakhir}2")
    ws["A2"] = NAMA_LLDIKTI
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = center

    for baris_num, teks in enumerate(
        [f"YAYASAN : {NAMA_YAYASAN}", f"PTS : {NAMA_PTS}", f"BULAN : {NAMA_BULAN[bulan - 1].upper()} {tahun}"],
        start=3,
    ):
        ws.merge_cells(f"A{baris_num}:{kolom_terakhir}{baris_num}")
        cell = ws[f"A{baris_num}"]
        cell.value = teks
        cell.font = Font(size=10)
        cell.alignment = center

    row_header = 6
    headers = ["No", "Nama/NIDN", "Gol/P Akademik"] + [str(t.day) for t, _ in jenis_list] + ["Total Hadir"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 22

    for idx, baris in enumerate(daftar, 1):
        row_num = row_header + idx
        ws.row_dimensions[row_num].height = 20

        ws.cell(row=row_num, column=1, value=idx).alignment = center
        nama_cell = ws.cell(
            row=row_num, column=2, value=f"{baris.user.get_full_name() or baris.user.username}\n{baris.user.nidn or ''}",
        )
        nama_cell.alignment = left
        ws.cell(row=row_num, column=3, value=baris.gol_jabatan).alignment = center

        if baris.tugas_belajar:
            kolom_awal = get_column_letter(4)
            kolom_akhir = get_column_letter(3 + jumlah_hari)
            ws.merge_cells(f"{kolom_awal}{row_num}:{kolom_akhir}{row_num}")
            sel = ws.cell(row=row_num, column=4, value="TUGAS BELAJAR / BIAYA PEMERINTAH")
            sel.alignment = center
            sel.font = Font(bold=True, size=9)
        else:
            for i, hari in enumerate(baris.hari_grid):
                col = 4 + i
                if hari.jenis == "kerja" and hari.hadir and baris.paraf:
                    gambar = _gambar_aman(baris.paraf.gambar)
                    if gambar:
                        ws.add_image(gambar, f"{get_column_letter(col)}{row_num}")

        total_cell = ws.cell(row=row_num, column=3 + jumlah_hari + 1, value="" if baris.tugas_belajar else baris.total_hadir)
        total_cell.alignment = center

        for col in range(1, jumlah_kolom + 1):
            ws.cell(row=row_num, column=col).border = thin

    # Arsir kolom Minggu/hari libur (semua baris data)
    baris_terakhir = row_header + len(daftar)
    for i, (_tanggal, jenis) in enumerate(jenis_list):
        if jenis != "kerja" and daftar:
            col = 4 + i
            for row_num in range(row_header + 1, baris_terakhir + 1):
                ws.cell(row=row_num, column=col).fill = arsir_fill

    col_widths = [5, 26, 12] + [4] * jumlah_hari + [10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # ---- Legenda (KET) ----
    baris_ket = baris_terakhir + 2
    ws.cell(row=baris_ket, column=1, value="KET:").font = Font(bold=True, size=9)
    ws.cell(row=baris_ket + 1, column=1).fill = arsir_fill
    ws.cell(row=baris_ket + 1, column=2, value=": Hari Minggu").font = Font(size=9)

    baris_ket_n = baris_ket + 2
    for libur in HariLibur.objects.filter(tanggal__year=tahun, tanggal__month=bulan).order_by("tanggal"):
        ws.cell(row=baris_ket_n, column=1).fill = arsir_fill
        teks = f": {libur.keterangan} ({libur.tanggal.strftime('%d %B %Y')})"
        ws.cell(row=baris_ket_n, column=2, value=teks).font = Font(size=9)
        baris_ket_n += 1

    # ---- Blok tanda tangan ----
    baris_ttd = baris_ket_n + 2
    kolom_ttd_letter = get_column_letter(max(jumlah_kolom - 3, 4))

    ws[f"{kolom_ttd_letter}{baris_ttd}"] = f"{kota}, {tanggal_cetak.strftime('%d %B %Y').upper()}"
    ws[f"{kolom_ttd_letter}{baris_ttd}"].alignment = center
    ws[f"{kolom_ttd_letter}{baris_ttd + 1}"] = f"{jabatan_penandatangan.upper()},"
    ws[f"{kolom_ttd_letter}{baris_ttd + 1}"].alignment = center

    baris_nama_ttd = baris_ttd + 6
    gambar_ttd = _gambar_aman(file_ttd_penandatangan) if file_ttd_penandatangan else None
    if gambar_ttd:
        ws.add_image(gambar_ttd, f"{kolom_ttd_letter}{baris_ttd + 2}")

    nama_cell = ws[f"{kolom_ttd_letter}{baris_nama_ttd}"]
    nama_cell.value = nama_penandatangan or "........................."
    nama_cell.font = Font(underline="single")
    nama_cell.alignment = center
    if nip_penandatangan:
        nip_cell = ws[f"{kolom_ttd_letter}{baris_nama_ttd + 1}"]
        nip_cell.value = f"NIP. {nip_penandatangan}"
        nip_cell.alignment = center

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
