from datetime import date, datetime

import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import HttpResponse, HttpResponseForbidden
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle
from rest_framework.views import APIView

from accounts.models import User
from laporan.views import get_dosen_queryset
from simda_dosen.utils import get_pejabat_aktif
from .decision import (
    cek_lokasi, hitung_ketepatan_masuk, hitung_ketepatan_pulang, resolve_kelompok, tentukan_status_waktu,
    verifikasi_wajah,
)
from .face import VERSI_MODEL_WAJAH, ekstrak_satu_wajah, enkripsi_embedding, rata_rata_embedding
from .forms import (
    HariLiburForm, IzinCutiForm, KelompokPresensiForm, LaporanInternalForm, LaporanSerdosForm,
    TargetKerjaBulananForm,
)
from .laporan_detail import cari_pegawai, detail_presensi_bulanan
from .laporan_detail_excel import render_excel_detail_presensi
from .laporan_detail_pdf import render_pdf_detail_presensi
from .laporan_internal import get_user_qs_laporan_internal
from .laporan_internal_excel import render_excel_laporan_internal
from .laporan_internal_pdf import render_pdf_laporan_internal
from .laporan_serdos import data_laporan_serdos
from .laporan_serdos_excel import render_excel_laporan_serdos
from .laporan_serdos_pdf import render_pdf_laporan_serdos
from .models import (
    BATAS_MENIT_LEMBUR_WAJIB_KETERANGAN, EnrolmentWajah, HariLibur, IzinCuti, KelompokPresensi, LogKecurangan,
    ParafDosen, Perangkat, Presensi, StatusApprovalLembur, StatusPresensi, TargetKerjaBulanan, TingkatRisiko,
    UrutanSerdos,
)
from .rekap import (
    data_presensi_harian, laporan_bulanan_jam_kerja, rekap_bulanan_user, ringkasan_hari_ini, top_telat_hari_ini,
    tren_mingguan,
)
from .serializers import AbsenSerializer, EnrolmentWajahSerializer, ParafSerializer
from .utils import get_dosen_serdos_qs

# Sama seperti accounts/views.py::ROLE_PENGELOLA_SCOPED -- peran yang boleh
# mengelola dosen dalam cakupannya sendiri (fakultas/prodi), dipakai di sini
# untuk menentukan siapa yang boleh meninjau presensi ditandai.
ROLE_PENGELOLA_SCOPED = ("dekan", "wadek", "kaprodi", "sekprodi", "operator")


def _sapaan_waktu(jam):
    if jam < 11:
        return "Selamat pagi"
    if jam < 15:
        return "Selamat siang"
    if jam < 19:
        return "Selamat sore"
    return "Selamat malam"


def _inisial_nama(nama):
    bagian = nama.split()
    if len(bagian) > 1:
        return (bagian[0][0] + bagian[-1][0]).upper()
    return nama[:2].upper() if nama else "?"


@login_required
def halaman_absen(request):
    """Halaman web untuk absen masuk/pulang (Tabler UI + JS geolocation,
    memanggil API /api/presensi/masuk & /pulang lewat sesi Django yang sudah
    login -- lihat SessionAuthentication di REST_FRAMEWORK settings)."""
    sudah_enrolment = EnrolmentWajah.objects.filter(user=request.user, consent_disetujui=True).exists()
    nama = request.user.get_full_name() or request.user.username
    return render(request, "presensi/absen.html", {
        "sudah_enrolment": sudah_enrolment,
        "sapaan": _sapaan_waktu(timezone.localtime(timezone.now()).hour),
        "nama_tampil": nama,
        "inisial": _inisial_nama(nama),
    })


@login_required
def halaman_enrolment(request):
    """Halaman web pendaftaran wajah (sekali di awal) -- ambil 2-5 selfie
    lewat kamera browser + persetujuan (consent), kirim ke
    /api/presensi/enrolment-wajah lewat sesi Django yang sudah login."""
    return render(request, "presensi/enrolment.html")


@login_required
def halaman_paraf(request):
    """Halaman web pembuatan paraf digital (canvas signature pad) --
    dasar bukti tanda tangan kehadiran untuk laporan ke LLDIKTI (dosen
    serdos). Lihat presensi/models.py::ParafDosen."""
    paraf = ParafDosen.objects.filter(user=request.user).first()
    return render(request, "presensi/paraf.html", {"paraf": paraf})


# Service worker minimal, sengaja TANPA caching berat -- halaman presensi
# butuh GPS/kamera live saat itu juga, jadi presensi offline tidak masuk
# akal. Isinya cuma cukup supaya browser menganggap /presensi/ "installable"
# (PWA Add to Home Screen).
_SERVICE_WORKER_JS = """
self.addEventListener('install', () => self.skipWaiting());
self.addEventListener('activate', (event) => event.waitUntil(self.clients.claim()));
self.addEventListener('fetch', () => {});
""".strip()


def service_worker_presensi(request):
    """Sengaja disajikan lewat view (BUKAN file statis WhiteNoise): service
    worker cuma boleh mengontrol path yang sama atau di bawah lokasi
    skripnya sendiri, kecuali server mengirim header Service-Worker-Allowed.
    Dengan disajikan di /presensi/sw.js, cakupannya otomatis mencakup
    seluruh /presensi/ tanpa perlu konfigurasi header tambahan di Nginx."""
    return HttpResponse(_SERVICE_WORKER_JS, content_type="application/javascript")


def _bisa_tinjau_presensi(user):
    return user.role == "admin" or user.role in ROLE_PENGELOLA_SCOPED


def _bisa_kelola_pengaturan_presensi(user):
    # BEDA dengan _bisa_tinjau_presensi (di-scope fakultas/prodi) -- jam
    # kerja/hari libur/target bulanan berlaku institusi-wide, jadi sengaja
    # dibatasi admin saja (pola sama dengan Kelola User/Data Master).
    return user.role == "admin"


def _tanggal_dari_request(request):
    tanggal_str = request.GET.get("tanggal")
    if tanggal_str:
        try:
            return datetime.strptime(tanggal_str, "%Y-%m-%d").date()
        except ValueError:
            pass
    return timezone.localdate()


@login_required
def dashboard_presensi(request):
    """Dashboard admin/HR: KPI hari ini, tren 6 hari, dan daftar paling
    telat. Cakupan DOSEN SAJA untuk sekarang (lewat get_dosen_queryset) --
    skema presensi sendiri sudah mendukung staf, lihat presensi/rekap.py."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    dosen_qs = get_dosen_queryset(request.user)
    user_ids = list(dosen_qs.values_list("id", flat=True))

    ringkasan = ringkasan_hari_ini(user_ids)
    tren = tren_mingguan(user_ids)
    top_telat = top_telat_hari_ini(user_ids)

    return render(request, "presensi/dashboard.html", {
        "ringkasan": ringkasan,
        "tren_labels": [t["label"] for t in tren],
        "tren_data": [t["jumlah"] for t in tren],
        "top_telat": top_telat,
    })


@login_required
def data_presensi(request):
    """Tabel data presensi harian (1 baris per dosen dalam cakupan,
    termasuk yang belum absen), dengan filter tanggal/fakultas/prodi."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    tanggal = _tanggal_dari_request(request)
    filter_fakultas = request.GET.get("fakultas", "")
    filter_prodi = request.GET.get("prodi", "")

    dosen_qs = get_dosen_queryset(request.user, filter_prodi=filter_prodi, filter_fakultas=filter_fakultas)
    daftar = data_presensi_harian(dosen_qs, tanggal)

    halaman = Paginator(daftar, 25).get_page(request.GET.get("halaman"))

    return render(request, "presensi/data.html", {
        "halaman": halaman,
        "tanggal": tanggal,
        "filter_fakultas": filter_fakultas,
        "filter_prodi": filter_prodi,
    })


@login_required
def export_excel_presensi(request):
    """Ekspor Excel data presensi harian -- style & pola sama dengan
    laporan/views.py supaya konsisten dengan ekspor lain di portal ini."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    tanggal = _tanggal_dari_request(request)
    filter_fakultas = request.GET.get("fakultas", "")
    filter_prodi = request.GET.get("prodi", "")

    dosen_qs = get_dosen_queryset(request.user, filter_prodi=filter_prodi, filter_fakultas=filter_fakultas)
    daftar = data_presensi_harian(dosen_qs, tanggal)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Presensi"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:J1")
    ws["A1"] = "DATA PRESENSI HARIAN"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:J2")
    ws["A2"] = f"Universitas Ichsan Gorontalo · {tanggal.strftime('%d %B %Y')}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = center

    headers = ["No", "Nama", "NIDN", "Fakultas", "Prodi", "Kelompok", "Masuk", "Pulang", "Status", "Lembur"]
    row_header = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 26

    for idx, item in enumerate(daftar, 1):
        dosen = item["dosen"]
        p = item["presensi"]
        row_data = [
            idx,
            dosen.get_full_name() or dosen.username,
            dosen.nidn,
            dosen.kode_fakultas or "-",
            dosen.kode_prodi or "-",
            p.kelompok.nama if p and p.kelompok else "-",
            timezone.localtime(p.waktu_masuk).strftime("%H:%M") if p and p.waktu_masuk else "-",
            timezone.localtime(p.waktu_pulang).strftime("%H:%M") if p and p.waktu_pulang else "-",
            p.get_status_display() if p else "Belum Absen",
            p.get_status_lembur_display() if p and p.menit_lembur else "-",
        ]
        row_num = row_header + idx
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            cell.alignment = center if col != 2 else left

    col_widths = [5, 28, 15, 10, 8, 14, 10, 10, 14, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Presensi_{tanggal.isoformat()}.xlsx"'
    wb.save(response)
    return response


def _bulan_tahun_dari_request(request):
    hari_ini = timezone.localdate()
    try:
        return int(request.GET.get("bulan", hari_ini.month)), int(request.GET.get("tahun", hari_ini.year))
    except ValueError:
        return hari_ini.month, hari_ini.year


def _pengguna_dalam_cakupan(request_user):
    """SEMUA role (dosen + staf/tendik + pejabat), BEDA dengan get_dosen_
    queryset (dosen-only) yang dipakai dashboard/data harian -- laporan
    bulanan ini memang dirancang lintas-pegawai sejak awal. Discope lewat
    dapat_kelola, pola sama dengan tinjau_presensi/tinjau_izin."""
    aktif = User.objects.filter(status_akun="aktif")
    ids_dalam_cakupan = [u.id for u in aktif if request_user.dapat_kelola(u)]
    return User.objects.filter(id__in=ids_dalam_cakupan)


@login_required
def laporan_bulanan_presensi(request):
    """Laporan rekap jam kerja BULANAN lintas-pegawai (dosen + staf) --
    dasar penggajian, lihat CLAUDE.md § 9. Beda dari dashboard_presensi/
    data_presensi yang masih dosen-only."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    bulan, tahun = _bulan_tahun_dari_request(request)
    daftar = laporan_bulanan_jam_kerja(_pengguna_dalam_cakupan(request.user), bulan, tahun)
    halaman = Paginator(daftar, 25).get_page(request.GET.get("halaman"))
    bulan_pilihan = [(i, datetime(2000, i, 1).strftime("%B")) for i in range(1, 13)]

    return render(request, "presensi/laporan_bulanan.html", {
        "halaman": halaman, "bulan": bulan, "tahun": tahun, "bulan_pilihan": bulan_pilihan,
    })


@login_required
def export_excel_laporan_bulanan(request):
    """Ekspor Excel laporan bulanan jam kerja -- gaya sama dengan
    export_excel_presensi."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    bulan, tahun = _bulan_tahun_dari_request(request)
    daftar = laporan_bulanan_jam_kerja(_pengguna_dalam_cakupan(request.user), bulan, tahun)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Laporan Bulanan"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    nama_bulan = datetime(tahun, bulan, 1).strftime("%B %Y")

    ws.merge_cells("A1:H1")
    ws["A1"] = "LAPORAN BULANAN JAM KERJA"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Universitas Ichsan Gorontalo · {nama_bulan}"
    ws["A2"].font = Font(bold=True, size=12)
    ws["A2"].alignment = center

    headers = ["No", "Nama", "Role", "Kelompok", "Hari Hadir", "Total Jam Kerja", "Target Jam Kerja", "Selisih"]
    row_header = 4
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_header].height = 26

    for idx, item in enumerate(daftar, 1):
        u = item["user"]
        target = item["target"]
        row_data = [
            idx,
            u.get_full_name() or u.username,
            u.get_role_display_id(),
            item["kelompok"].nama if item["kelompok"] else "-",
            item["hari_hadir"],
            item["total_jam_kerja"],
            f"{target.target_jam_kerja} jam" if target else "-",
            item["selisih_jam_kerja"] or "-",
        ]
        row_num = row_header + idx
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            cell.alignment = center if col != 2 else left

    col_widths = [5, 28, 16, 16, 10, 14, 14, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Laporan_Bulanan_{tahun}-{bulan:02d}.xlsx"'
    wb.save(response)
    return response


def _dapatkan_user_qs_dari_form_internal(request, data):
    """Cakupan (dapat_kelola) + filter kategori/fakultas/prodi -- dipakai
    bareng oleh halaman & kedua ekspor Laporan Presensi Internal."""
    return get_user_qs_laporan_internal(
        _pengguna_dalam_cakupan(request.user),
        kategori=data.get("kategori", ""), fakultas=data.get("fakultas", ""), prodi=data.get("prodi", ""),
    )


@login_required
def halaman_laporan_internal(request):
    """Laporan Presensi Internal (BEDA dari Laporan Bulanan Jam Kerja --
    ini grid tanggal dengan jam Masuk/Pulang per hari, bukan agregat
    total jam kerja). Sama seperti Laporan Bulanan: cakupan lintas-
    pegawai (dosen+staf+pejabat) di-scope lewat dapat_kelola, BUKAN
    admin-only."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    if "bulan" in request.GET:
        form = LaporanInternalForm(request.GET)
    else:
        hari_ini = timezone.localdate()
        form = LaporanInternalForm(initial={"bulan": hari_ini.month, "tahun": hari_ini.year})

    return render(request, "presensi/laporan_internal.html", {"form": form})


@login_required
def export_pdf_laporan_internal(request):
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    form = LaporanInternalForm(request.GET)
    if not form.is_valid():
        messages.error(request, "Parameter laporan tidak valid, periksa kembali isian form.")
        return redirect("presensi_web:laporan_internal")

    data = form.cleaned_data
    bulan, tahun = int(data["bulan"]), data["tahun"]
    user_qs = _dapatkan_user_qs_dari_form_internal(request, data)

    pdf_bytes = render_pdf_laporan_internal(
        user_qs, bulan, tahun, kategori=data["kategori"], fakultas=data["fakultas"], prodi=data["prodi"],
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Laporan_Presensi_Internal_{tahun}-{bulan:02d}.pdf"'
    return response


@login_required
def export_excel_laporan_internal(request):
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    form = LaporanInternalForm(request.GET)
    if not form.is_valid():
        messages.error(request, "Parameter laporan tidak valid, periksa kembali isian form.")
        return redirect("presensi_web:laporan_internal")

    data = form.cleaned_data
    bulan, tahun = int(data["bulan"]), data["tahun"]
    user_qs = _dapatkan_user_qs_dari_form_internal(request, data)

    buffer = render_excel_laporan_internal(
        user_qs, bulan, tahun, kategori=data["kategori"], fakultas=data["fakultas"], prodi=data["prodi"],
    )
    response = HttpResponse(
        buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="Laporan_Presensi_Internal_{tahun}-{bulan:02d}.xlsx"'
    return response


def _cari_target_user_laporan_detail(request):
    """Tentukan target_user (kalau ada) + daftar_pilihan (kalau pencarian
    cocok ke lebih dari 1 orang) untuk Detail Presensi Bulanan -- dipakai
    bareng oleh halaman & kedua ekspor."""
    cakupan = _pengguna_dalam_cakupan(request.user)
    user_id = request.GET.get("user_id", "").strip()
    if user_id.isdigit():
        return get_object_or_404(cakupan, id=int(user_id)), None

    kata_kunci = request.GET.get("q", "").strip()
    if not kata_kunci:
        return None, None

    hasil = cari_pegawai(cakupan, kata_kunci)
    jumlah = hasil.count()
    if jumlah == 1:
        return hasil.first(), None
    if jumlah > 1:
        return None, hasil
    return None, None


@login_required
def halaman_laporan_detail(request):
    """Detail Presensi Bulanan -- fitur OPSIONAL untuk drill-down 1 orang
    tertentu (dicari lewat nama/NIDN), BEDA dari 3 laporan presensi lain
    yang selalu menampilkan banyak orang sekaligus. Alur: cari -> kalau
    cocok lebih dari 1, pilih dari daftar -> tampil detail + ringkasan
    (reuse rekap_bulanan_user, lihat presensi/laporan_detail.py)."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    kata_kunci = request.GET.get("q", "").strip()
    bulan, tahun = _bulan_tahun_dari_request(request)
    target_user, daftar_pilihan = _cari_target_user_laporan_detail(request)

    detail = detail_presensi_bulanan(target_user, bulan, tahun) if target_user else None

    return render(request, "presensi/laporan_detail.html", {
        "kata_kunci": kata_kunci, "bulan": bulan, "tahun": tahun,
        "target_user": target_user, "daftar_pilihan": daftar_pilihan, "detail": detail,
    })


@login_required
def export_pdf_laporan_detail(request):
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    target_user, _daftar_pilihan = _cari_target_user_laporan_detail(request)
    if not target_user:
        messages.error(request, "Pilih pegawai dulu sebelum mengunduh.")
        return redirect("presensi_web:laporan_detail")

    bulan, tahun = _bulan_tahun_dari_request(request)
    pdf_bytes = render_pdf_detail_presensi(target_user, bulan, tahun)
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = (
        f'attachment; filename="Detail_Presensi_{target_user.username}_{tahun}-{bulan:02d}.pdf"'
    )
    return response


@login_required
def export_excel_laporan_detail(request):
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    target_user, _daftar_pilihan = _cari_target_user_laporan_detail(request)
    if not target_user:
        messages.error(request, "Pilih pegawai dulu sebelum mengunduh.")
        return redirect("presensi_web:laporan_detail")

    bulan, tahun = _bulan_tahun_dari_request(request)
    buffer = render_excel_detail_presensi(target_user, bulan, tahun)
    response = HttpResponse(
        buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="Detail_Presensi_{target_user.username}_{tahun}-{bulan:02d}.xlsx"'
    )
    return response


@login_required
def tinjau_presensi(request):
    """Halaman HR/admin: daftar presensi yang ditandai (tingkat_risiko
    sedang/tinggi) untuk ditinjau manual -- lihat CLAUDE.md § 3. Dibatasi ke
    orang dalam cakupan reviewer (fakultas/prodi) lewat
    accounts.User.dapat_kelola, sama seperti pola yang sudah dipakai app
    profil (dulu lewat wrapper dapat_kelola_nidn -- sekarang bisa langsung
    karena Presensi.user adalah FK asli ke accounts.User)."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    semua_ditandai = (
        Presensi.objects.filter(ditandai=True)
        .select_related("lokasi", "user")
        .order_by("-tanggal", "user__username")
    )
    antrian = [p for p in semua_ditandai if request.user.dapat_kelola(p.user)]

    daftar = [{"presensi": p, "nama_dosen": p.user.get_full_name() or p.user.username} for p in antrian]

    semua_lembur = (
        Presensi.objects.filter(status_lembur=StatusApprovalLembur.MENUNGGU)
        .select_related("lokasi", "user")
        .order_by("-tanggal", "user__username")
    )
    antrian_lembur = [p for p in semua_lembur if request.user.dapat_kelola(p.user)]
    daftar_lembur = [
        {"presensi": p, "nama_dosen": p.user.get_full_name() or p.user.username} for p in antrian_lembur
    ]

    return render(request, "presensi/tinjau.html", {"daftar": daftar, "daftar_lembur": daftar_lembur})


@login_required
def putuskan_presensi(request, presensi_id):
    """POST-only: HR/admin menyetujui atau menolak satu presensi ditandai."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method != "POST":
        return redirect("presensi_web:tinjau")

    presensi = get_object_or_404(Presensi.objects.select_related("user"), id=presensi_id, ditandai=True)
    if not request.user.dapat_kelola(presensi.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke presensi orang ini.")

    aksi = request.POST.get("aksi")
    if aksi == "setujui":
        presensi.ditandai = False
        presensi.tingkat_risiko = TingkatRisiko.RENDAH
        presensi.save(update_fields=["ditandai", "tingkat_risiko"])
        messages.success(request, f"Presensi {presensi.user} tanggal {presensi.tanggal} disetujui.")
    elif aksi == "tolak":
        presensi.status = StatusPresensi.DITOLAK
        presensi.ditandai = False
        presensi.save(update_fields=["status", "ditandai"])
        LogKecurangan.objects.create(
            user=presensi.user, presensi=presensi, jenis_anomali="ditolak_hr",
            skor=100, detail={"oleh": request.user.username},
        )
        messages.warning(request, f"Presensi {presensi.user} tanggal {presensi.tanggal} ditolak.")
    else:
        messages.error(request, "Aksi tidak dikenal.")

    return redirect("presensi_web:tinjau")


@login_required
def putuskan_lembur(request, presensi_id):
    """POST-only: atasan menyetujui/menolak keterangan lembur satu presensi.
    Kalau ditolak, jam pulang efektif (Presensi.durasi_kerja) otomatis
    kembali dibatasi ke jam pulang normal kelompok/lokasi -- lihat
    Presensi.durasi_kerja_menit -- waktu_pulang ASLI tetap tersimpan apa
    adanya, tidak diubah."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method != "POST":
        return redirect("presensi_web:tinjau")

    presensi = get_object_or_404(
        Presensi.objects.select_related("user"), id=presensi_id, status_lembur=StatusApprovalLembur.MENUNGGU,
    )
    if not request.user.dapat_kelola(presensi.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke presensi orang ini.")

    aksi = request.POST.get("aksi")
    if aksi == "setujui":
        presensi.status_lembur = StatusApprovalLembur.DISETUJUI
    elif aksi == "tolak":
        presensi.status_lembur = StatusApprovalLembur.DITOLAK
    else:
        messages.error(request, "Aksi tidak dikenal.")
        return redirect("presensi_web:tinjau")

    presensi.approver_lembur = request.user
    presensi.waktu_keputusan_lembur = timezone.now()
    presensi.save(update_fields=["status_lembur", "approver_lembur", "waktu_keputusan_lembur"])

    if aksi == "setujui":
        messages.success(request, f"Lembur {presensi.user} tanggal {presensi.tanggal} disetujui.")
    else:
        messages.warning(request, f"Lembur {presensi.user} tanggal {presensi.tanggal} ditolak.")

    return redirect("presensi_web:tinjau")


@login_required
def halaman_riwayat(request):
    """Riwayat presensi PRIBADI (bulan berjalan) -- gabungan Presensi &
    IzinCuti yang disetujui, diurutkan dari yang terbaru."""
    hari_ini = timezone.localdate()
    awal_bulan = hari_ini.replace(day=1)

    presensi_list = Presensi.objects.filter(user=request.user, tanggal__gte=awal_bulan).select_related("lokasi")
    izin_list = IzinCuti.objects.filter(
        user=request.user, status=IzinCuti.StatusApproval.DISETUJUI, tanggal_mulai__gte=awal_bulan,
    )

    entri = [{"jenis": "presensi", "tanggal": p.tanggal, "obj": p} for p in presensi_list]
    entri += [{"jenis": "izin", "tanggal": i.tanggal_mulai, "obj": i} for i in izin_list]
    entri.sort(key=lambda e: e["tanggal"], reverse=True)

    rekap_bulan = rekap_bulanan_user(request.user, hari_ini.month, hari_ini.year)

    return render(request, "presensi/riwayat.html", {
        "entri": entri,
        "bulan_label": hari_ini.strftime("%B %Y"),
        "rekap_bulan": rekap_bulan,
    })


@login_required
def halaman_izin(request):
    """Pengajuan izin/sakit/cuti/dinas mandiri + riwayat pengajuan pribadi."""
    if request.method == "POST":
        form = IzinCutiForm(request.POST, request.FILES)
        if form.is_valid():
            izin = form.save(commit=False)
            izin.user = request.user
            izin.save()
            messages.success(request, "Pengajuan izin berhasil dikirim ke atasan.")
            return redirect("presensi_web:izin")
        messages.error(request, "Periksa kembali form pengajuan Anda.")
    else:
        form = IzinCutiForm()

    riwayat = IzinCuti.objects.filter(user=request.user).order_by("-dibuat")[:20]
    return render(request, "presensi/izin.html", {"form": form, "riwayat": riwayat})


@login_required
def tinjau_izin(request):
    """Halaman atasan: daftar pengajuan izin yang menunggu persetujuan,
    di-scope lewat accounts.User.dapat_kelola -- sama seperti tinjau_presensi."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    menunggu = (
        IzinCuti.objects.filter(status=IzinCuti.StatusApproval.MENUNGGU)
        .select_related("user")
        .order_by("tanggal_mulai")
    )
    daftar = [i for i in menunggu if request.user.dapat_kelola(i.user)]
    return render(request, "presensi/izin_tinjau.html", {"daftar": daftar})


@login_required
def putuskan_izin(request, izin_id):
    """POST-only: atasan menyetujui/menolak satu pengajuan izin."""
    if not _bisa_tinjau_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method != "POST":
        return redirect("presensi_web:izin_tinjau")

    izin = get_object_or_404(
        IzinCuti.objects.select_related("user"), id=izin_id, status=IzinCuti.StatusApproval.MENUNGGU,
    )
    if not request.user.dapat_kelola(izin.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke pengajuan orang ini.")

    aksi = request.POST.get("aksi")
    if aksi == "setujui":
        izin.status = IzinCuti.StatusApproval.DISETUJUI
        izin.approver = request.user
        izin.save(update_fields=["status", "approver"])
        messages.success(request, f"Pengajuan izin {izin.user} disetujui.")
    elif aksi == "tolak":
        izin.status = IzinCuti.StatusApproval.DITOLAK
        izin.approver = request.user
        izin.save(update_fields=["status", "approver"])
        messages.warning(request, f"Pengajuan izin {izin.user} ditolak.")
    else:
        messages.error(request, "Aksi tidak dikenal.")

    return redirect("presensi_web:izin_tinjau")


# Skor risiko presensi yang lolos KEDUA syarat (lokasi & wajah) -- rendah
# tapi tidak nol, sisakan ruang untuk sinyal tambahan (QR/Wi-Fi) nanti.
SKOR_RISIKO_TERVERIFIKASI = 8

# Skor risiko dicatat ke LogKecurangan per jenis alasan gagal. Alasan yang
# TIDAK ada di sini (mis. "belum_enrolment_wajah", "foto_tidak_valid") sengaja
# tidak dicatat sebagai kecurangan -- itu soal kesiapan data, bukan indikasi
# curang.
SKOR_ANOMALI = {
    "akurasi_buruk": 60,
    "di_luar_radius": 70,
    "liveness_gagal": 75,
    "wajah_tidak_cocok": 90,
}


class PresensiRateThrottle(UserRateThrottle):
    scope = "presensi"


def _catat_perangkat(user, device_id, waktu):
    Perangkat.objects.update_or_create(
        user=user, device_id=device_id,
        defaults={"terakhir_dipakai": waktu},
    )


def _catat_jika_anomali(user, presensi, alasan, detail):
    if alasan in SKOR_ANOMALI:
        LogKecurangan.objects.create(
            user=user, presensi=presensi, jenis_anomali=alasan,
            skor=SKOR_ANOMALI[alasan], detail=detail,
        )


def _respon_ditolak(alasan):
    return Response({"diterima": False, "alasan": alasan, "tingkat_risiko": TingkatRisiko.TINGGI})


def _jalankan_gerbang(user, data, presensi=None):
    """Gerbang-DAN: cek_lokasi lalu (kalau lolos) verifikasi_wajah. Return
    (hasil_lokasi, None) kalau dua-duanya lolos, atau (None, Response) kalau
    salah satu gagal -- lihat presensi/decision.py."""
    hasil_lokasi = cek_lokasi(data["lat"], data["lng"], data["akurasi_m"])
    if not hasil_lokasi.lolos:
        _catat_jika_anomali(
            user, presensi, hasil_lokasi.alasan,
            {"lat": data["lat"], "lng": data["lng"], "akurasi_m": data["akurasi_m"]},
        )
        return None, _respon_ditolak(hasil_lokasi.alasan)

    hasil_wajah = verifikasi_wajah(user, data["selfie"])
    if not hasil_wajah.lolos:
        _catat_jika_anomali(user, presensi, hasil_wajah.alasan, {"skor_kemiripan": hasil_wajah.skor_kemiripan})
        return None, _respon_ditolak(hasil_wajah.alasan)

    return hasil_lokasi, None


class AbsenMasukView(APIView):
    """POST /api/presensi/masuk — absen masuk, gerbang-DAN cek lokasi + wajah."""
    throttle_classes = [PresensiRateThrottle]

    def post(self, request):
        serializer = AbsenSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        user = request.user

        waktu_server = timezone.localtime(timezone.now())
        tanggal = waktu_server.date()

        if Presensi.objects.filter(user=user, tanggal=tanggal, waktu_masuk__isnull=False).exists():
            return Response(
                {"diterima": False, "alasan": "sudah_absen_masuk"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        _catat_perangkat(user, data["device_id"], waktu_server)

        hasil_lokasi, response_gagal = _jalankan_gerbang(user, data)
        if response_gagal is not None:
            return response_gagal

        kelompok = resolve_kelompok(user)
        status_kehadiran = tentukan_status_waktu(
            hasil_lokasi.lokasi, waktu_server.time(), kelompok=kelompok, tanggal=tanggal,
        )
        menit_lebih_awal, menit_terlambat = hitung_ketepatan_masuk(
            hasil_lokasi.lokasi, waktu_server.time(), kelompok=kelompok, tanggal=tanggal,
        )

        presensi, _ = Presensi.objects.update_or_create(
            user=user, tanggal=tanggal,
            defaults={
                "waktu_masuk": waktu_server,
                "lokasi": hasil_lokasi.lokasi,
                # Snapshot kelompok yang berlaku SAAT ini -- lihat catatan di
                # KelompokPresensi soal kenapa bukan FK live ke role user.
                "kelompok": kelompok,
                "latitude_masuk": data["lat"],
                "longitude_masuk": data["lng"],
                "akurasi_masuk_m": data["akurasi_m"],
                "status": status_kehadiran,
                "menit_lebih_awal": menit_lebih_awal,
                "menit_terlambat": menit_terlambat,
                "skor_risiko": SKOR_RISIKO_TERVERIFIKASI,
                "tingkat_risiko": TingkatRisiko.RENDAH,
                "ditandai": False,
            },
        )

        return Response({
            "diterima": True,
            "status": presensi.status,
            "waktu_masuk": presensi.waktu_masuk,
            "lokasi": hasil_lokasi.lokasi.nama,
            "skor_risiko": presensi.skor_risiko,
            "tingkat_risiko": presensi.tingkat_risiko,
        })


class AbsenPulangView(APIView):
    """POST /api/presensi/pulang — absen pulang, gerbang-DAN cek lokasi + wajah."""
    throttle_classes = [PresensiRateThrottle]

    def post(self, request):
        serializer = AbsenSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data
        user = request.user

        waktu_server = timezone.localtime(timezone.now())
        tanggal = waktu_server.date()

        try:
            presensi = Presensi.objects.get(user=user, tanggal=tanggal)
        except Presensi.DoesNotExist:
            return Response(
                {"diterima": False, "alasan": "belum_absen_masuk"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if presensi.waktu_pulang:
            return Response(
                {"diterima": False, "alasan": "sudah_absen_pulang"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        _catat_perangkat(user, data["device_id"], waktu_server)

        hasil_lokasi, response_gagal = _jalankan_gerbang(user, data, presensi=presensi)
        if response_gagal is not None:
            return response_gagal

        # Pakai kelompok yang SUDAH di-snapshot saat absen masuk (bukan
        # resolve_kelompok(user) lagi) -- konsisten dengan jam yang jadi
        # acuan status kedatangan tadi pagi, lihat catatan di KelompokPresensi.
        # lokasi diambil dari presensi.lokasi kalau sudah ada (masuk normal),
        # fallback ke lokasi hasil gerbang pulang untuk baris lama yang belum
        # sempat punya lokasi tersimpan.
        lokasi_acuan = presensi.lokasi or hasil_lokasi.lokasi
        menit_pulang_cepat, menit_lembur = hitung_ketepatan_pulang(
            lokasi_acuan, waktu_server.time(), kelompok=presensi.kelompok, tanggal=tanggal,
        )
        keterangan_lembur = data["keterangan_lembur"].strip()
        if menit_lembur > BATAS_MENIT_LEMBUR_WAJIB_KETERANGAN and not keterangan_lembur:
            return Response({"diterima": False, "alasan": "keterangan_lembur_wajib"})

        presensi.waktu_pulang = waktu_server
        presensi.lokasi = lokasi_acuan
        presensi.latitude_pulang = data["lat"]
        presensi.longitude_pulang = data["lng"]
        presensi.menit_pulang_cepat = menit_pulang_cepat
        presensi.menit_lembur = menit_lembur
        if menit_lembur > BATAS_MENIT_LEMBUR_WAJIB_KETERANGAN:
            presensi.keterangan_lembur = keterangan_lembur
            presensi.status_lembur = StatusApprovalLembur.MENUNGGU
        presensi.save()

        return Response({
            "diterima": True,
            "status": presensi.status,
            "waktu_pulang": presensi.waktu_pulang,
            "lokasi": hasil_lokasi.lokasi.nama,
            "menit_lembur": presensi.menit_lembur,
            "perlu_persetujuan_lembur": presensi.status_lembur == StatusApprovalLembur.MENUNGGU,
        })


class EnrolmentWajahView(APIView):
    """POST /api/presensi/enrolment-wajah — daftarkan wajah (sekali di
    awal, sebelum bisa lolos syarat 2 saat absen)."""
    throttle_classes = [PresensiRateThrottle]

    def post(self, request):
        serializer = EnrolmentWajahSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        data = serializer.validated_data

        embeddings = []
        for foto in data["foto"]:
            wajah, _alasan = ekstrak_satu_wajah(foto)
            if wajah is not None:
                embeddings.append(wajah.embedding)

        if len(embeddings) < 2:
            return Response(
                {"status": "gagal", "alasan": "wajah_tidak_terdeteksi_konsisten"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        embedding_rata = rata_rata_embedding(embeddings)
        waktu = timezone.now()

        EnrolmentWajah.objects.update_or_create(
            user=request.user,
            defaults={
                "embedding_terenkripsi": enkripsi_embedding(embedding_rata),
                "versi_model": VERSI_MODEL_WAJAH,
                "consent_disetujui": True,
                "consent_pada": waktu,
            },
        )

        return Response({"status": "ok", "versi_model": VERSI_MODEL_WAJAH, "consent_pada": waktu})


class ParafDosenView(APIView):
    """POST /api/presensi/paraf — simpan/ganti paraf digital (gambar hasil
    canvas). Sekali gambar, dipakai berulang (auto-stamp) untuk laporan
    presensi bertanda-tangan ke LLDIKTI -- lihat presensi/models.py::
    ParafDosen."""
    throttle_classes = [PresensiRateThrottle]

    def post(self, request):
        serializer = ParafSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ParafDosen.objects.update_or_create(
            user=request.user, defaults={"gambar": serializer.validated_data["gambar"]},
        )
        return Response({"status": "ok"})


class StatusHariIniView(APIView):
    """GET /api/presensi/status-hari-ini — status presensi hari ini."""

    def get(self, request):
        presensi = Presensi.objects.filter(user=request.user, tanggal=timezone.localdate()).first()
        if not presensi:
            return Response({"sudah_absen_masuk": False, "sudah_absen_pulang": False})

        return Response({
            "sudah_absen_masuk": presensi.waktu_masuk is not None,
            "sudah_absen_pulang": presensi.waktu_pulang is not None,
            "status": presensi.status,
            "waktu_masuk": presensi.waktu_masuk,
            "waktu_pulang": presensi.waktu_pulang,
        })


# ============================================================
# Pengaturan Presensi (Kelompok, Hari Libur, Target Kerja Bulanan)
# -- admin-only, lihat _bisa_kelola_pengaturan_presensi. Sebelumnya cuma
# bisa diatur lewat Django Admin; halaman ini supaya HR/admin tidak perlu
# ke /admin/ untuk kerjaan rutin (tambah hari libur, atur target bulanan).
# ============================================================

@login_required
def pengaturan_kelompok(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    daftar = KelompokPresensi.objects.all().order_by("nama")
    return render(request, "presensi/pengaturan_kelompok.html", {"daftar": daftar})


@login_required
def tambah_kelompok(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method == "POST":
        form = KelompokPresensiForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Kelompok presensi berhasil ditambahkan.")
            return redirect("presensi_web:pengaturan_kelompok")
    else:
        form = KelompokPresensiForm()
    return render(request, "presensi/kelompok_form.html", {"form": form, "judul": "Tambah Kelompok Presensi"})


@login_required
def ubah_kelompok(request, kelompok_id):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    kelompok = get_object_or_404(KelompokPresensi, id=kelompok_id)
    if request.method == "POST":
        form = KelompokPresensiForm(request.POST, instance=kelompok)
        if form.is_valid():
            form.save()
            messages.success(request, "Kelompok presensi berhasil diperbarui.")
            return redirect("presensi_web:pengaturan_kelompok")
    else:
        form = KelompokPresensiForm(instance=kelompok)
    return render(request, "presensi/kelompok_form.html", {
        "form": form, "judul": f"Ubah Kelompok: {kelompok.nama}",
    })


@login_required
def toggle_aktif_kelompok(request, kelompok_id):
    """POST-only: aktifkan/nonaktifkan kelompok -- BUKAN hapus, karena
    KelompokPresensi direferensikan lewat Presensi.kelompok (on_delete=
    PROTECT, snapshot histori) sehingga tidak boleh dihapus kalau sudah
    dipakai. Field `aktif` sudah dirancang untuk kebutuhan ini."""
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method != "POST":
        return redirect("presensi_web:pengaturan_kelompok")
    kelompok = get_object_or_404(KelompokPresensi, id=kelompok_id)
    kelompok.aktif = not kelompok.aktif
    kelompok.save(update_fields=["aktif"])
    messages.success(request, f"Kelompok {kelompok.nama} sekarang {'aktif' if kelompok.aktif else 'nonaktif'}.")
    return redirect("presensi_web:pengaturan_kelompok")


@login_required
def pengaturan_hari_libur(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    daftar = HariLibur.objects.all().order_by("-tanggal")
    return render(request, "presensi/pengaturan_hari_libur.html", {"daftar": daftar})


@login_required
def tambah_hari_libur(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method == "POST":
        form = HariLiburForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Hari libur berhasil ditambahkan.")
            return redirect("presensi_web:pengaturan_hari_libur")
    else:
        form = HariLiburForm()
    return render(request, "presensi/hari_libur_form.html", {"form": form})


@login_required
def hapus_hari_libur(request, hari_libur_id):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method != "POST":
        return redirect("presensi_web:pengaturan_hari_libur")
    hari_libur = get_object_or_404(HariLibur, id=hari_libur_id)
    hari_libur.delete()
    messages.success(request, "Hari libur berhasil dihapus.")
    return redirect("presensi_web:pengaturan_hari_libur")


@login_required
def pengaturan_target(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    daftar = TargetKerjaBulanan.objects.select_related("kelompok").order_by("-tahun", "-bulan", "kelompok__nama")
    return render(request, "presensi/pengaturan_target.html", {"daftar": daftar})


@login_required
def tambah_target(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method == "POST":
        form = TargetKerjaBulananForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Target kerja bulanan berhasil ditambahkan.")
            return redirect("presensi_web:pengaturan_target")
    else:
        form = TargetKerjaBulananForm()
    return render(request, "presensi/target_form.html", {"form": form, "judul": "Tambah Target Kerja Bulanan"})


@login_required
def ubah_target(request, target_id):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    target = get_object_or_404(TargetKerjaBulanan, id=target_id)
    if request.method == "POST":
        form = TargetKerjaBulananForm(request.POST, instance=target)
        if form.is_valid():
            form.save()
            messages.success(request, "Target kerja bulanan berhasil diperbarui.")
            return redirect("presensi_web:pengaturan_target")
    else:
        form = TargetKerjaBulananForm(instance=target)
    return render(request, "presensi/target_form.html", {"form": form, "judul": "Ubah Target Kerja Bulanan"})


@login_required
def hapus_target(request, target_id):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")
    if request.method != "POST":
        return redirect("presensi_web:pengaturan_target")
    target = get_object_or_404(TargetKerjaBulanan, id=target_id)
    target.delete()
    messages.success(request, "Target kerja bulanan berhasil dihapus.")
    return redirect("presensi_web:pengaturan_target")


@login_required
def pengaturan_urutan_serdos(request):
    """Atur nomor urut dosen serdos untuk Laporan Daftar Hadir Dosen Serdos
    (LLDIKTI) -- urutan resmi kepegawaian tidak tersimpan di data mana
    pun, jadi diatur manual di sini. Satu form besar (bukan tambah/ubah
    per baris) supaya HR bisa isi semua sekaligus -- lihat
    presensi/utils.py::get_dosen_serdos_qs untuk cakupannya."""
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    dosen_serdos = list(get_dosen_serdos_qs().order_by("first_name", "username"))

    if request.method == "POST":
        for u in dosen_serdos:
            nilai = request.POST.get(f"urutan_{u.id}", "").strip()
            if nilai.isdigit():
                UrutanSerdos.objects.update_or_create(user=u, defaults={"urutan": int(nilai)})
            else:
                UrutanSerdos.objects.filter(user=u).delete()
        messages.success(request, "Urutan laporan serdos berhasil disimpan.")
        return redirect("presensi_web:pengaturan_urutan_serdos")

    urutan_map = {us.user_id: us.urutan for us in UrutanSerdos.objects.filter(user__in=dosen_serdos)}
    daftar = [{"user": u, "urutan": urutan_map.get(u.id)} for u in dosen_serdos]
    daftar.sort(key=lambda d: (d["urutan"] is None, d["urutan"] or 0))

    return render(request, "presensi/pengaturan_urutan_serdos.html", {"daftar": daftar})


def _default_parameter_laporan_serdos():
    """Nilai awal form Laporan Daftar Hadir Serdos: bulan/tahun berjalan,
    tanggal cetak default tanggal 1 bulan BERIKUTNYA (pola yang terlihat
    di contoh dokumen fisik -- laporan bulan berjalan ditandatangani awal
    bulan depan), dan nama/NIP penandatangan diambil otomatis dari
    Pejabat Struktural SIMDA kalau ketemu."""
    hari_ini = timezone.localdate()
    if hari_ini.month == 12:
        tanggal_cetak_default = date(hari_ini.year + 1, 1, 1)
    else:
        tanggal_cetak_default = date(hari_ini.year, hari_ini.month + 1, 1)

    pejabat = get_pejabat_aktif("Rektor")
    nama_default = pejabat.dosen.nama_lengkap_gelar if pejabat and pejabat.dosen else ""
    nip_default = pejabat.dosen.nip if pejabat and pejabat.dosen else ""

    return {
        "bulan": hari_ini.month, "tahun": hari_ini.year, "kota": "Gorontalo",
        "tanggal_cetak": tanggal_cetak_default, "jabatan_penandatangan": "Rektor",
        "nama_penandatangan": nama_default, "nip_penandatangan": nip_default,
    }


def _resolve_penandatangan(data):
    """Nama/NIP dari isian form (kalau diisi manual), fallback ke data
    Pejabat Struktural SIMDA untuk jabatan itu. Gambar tanda tangan
    SELALU ikut jabatan (bukan nama yang diketik) -- kalau HR cuma benerin
    typo gelar di form, tanda tangan yang benar tetap terpasang."""
    pejabat = get_pejabat_aktif(data["jabatan_penandatangan"])
    nama = data["nama_penandatangan"] or (pejabat.dosen.nama_lengkap_gelar if pejabat and pejabat.dosen else "")
    nip = data["nip_penandatangan"] or (pejabat.dosen.nip if pejabat and pejabat.dosen else "")
    file_ttd = pejabat.file_ttd if pejabat else None
    return nama, nip, file_ttd


@login_required
def halaman_laporan_serdos(request):
    """Halaman generate Laporan Daftar Hadir Dosen Serdos (LLDIKTI) --
    admin-only, sama seperti pengaturan lain. Tombol unduh PDF/Excel
    pakai formaction supaya satu form bisa kirim ke dua endpoint beda
    tergantung tombol yang diklik (lihat templates/presensi/
    laporan_serdos.html)."""
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    if "bulan" in request.GET:
        form = LaporanSerdosForm(request.GET)
    else:
        form = LaporanSerdosForm(initial=_default_parameter_laporan_serdos())

    return render(request, "presensi/laporan_serdos.html", {
        "form": form, "jumlah_dosen_serdos": get_dosen_serdos_qs().count(),
    })


@login_required
def export_pdf_daftar_hadir_serdos(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    form = LaporanSerdosForm(request.GET)
    if not form.is_valid():
        messages.error(request, "Parameter laporan tidak valid, periksa kembali isian form.")
        return redirect("presensi_web:laporan_serdos")

    data = form.cleaned_data
    bulan, tahun = int(data["bulan"]), data["tahun"]
    nama, nip, file_ttd = _resolve_penandatangan(data)

    pdf_bytes = render_pdf_laporan_serdos(
        bulan=bulan, tahun=tahun, kota=data["kota"], tanggal_cetak=data["tanggal_cetak"],
        jabatan_penandatangan=data["jabatan_penandatangan"], nama_penandatangan=nama,
        nip_penandatangan=nip, file_ttd_penandatangan=file_ttd,
    )
    response = HttpResponse(pdf_bytes, content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="Daftar_Hadir_Serdos_{tahun}-{bulan:02d}.pdf"'
    return response


@login_required
def export_excel_daftar_hadir_serdos(request):
    if not _bisa_kelola_pengaturan_presensi(request.user):
        return HttpResponseForbidden("Anda tidak memiliki akses ke halaman ini.")

    form = LaporanSerdosForm(request.GET)
    if not form.is_valid():
        messages.error(request, "Parameter laporan tidak valid, periksa kembali isian form.")
        return redirect("presensi_web:laporan_serdos")

    data = form.cleaned_data
    bulan, tahun = int(data["bulan"]), data["tahun"]
    nama, nip, file_ttd = _resolve_penandatangan(data)

    buffer = render_excel_laporan_serdos(
        bulan=bulan, tahun=tahun, kota=data["kota"], tanggal_cetak=data["tanggal_cetak"],
        jabatan_penandatangan=data["jabatan_penandatangan"], nama_penandatangan=nama,
        nip_penandatangan=nip, file_ttd_penandatangan=file_ttd,
    )
    response = HttpResponse(
        buffer.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="Daftar_Hadir_Serdos_{tahun}-{bulan:02d}.xlsx"'
    return response
