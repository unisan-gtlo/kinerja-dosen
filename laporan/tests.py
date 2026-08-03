import io
from unittest.mock import MagicMock, patch

import openpyxl
from django.test import TestCase, Client
from django.urls import reverse

from accounts.models import User
from master.models import Fakultas, Prodi
from laporan.views import _identitas_tampilan_dosen


class IdentitasTampilanDosenTest(TestCase):
    """laporan.views._identitas_tampilan_dosen -- nama+gelar dan
    fakultas/prodi tampilan laporan harus utamakan data SIMDA (profil),
    bukan data akun (accounts.User), konsisten dengan pola yang sudah
    dipakai dashboard/views.py::rekap(). Bug fix per 2026-08-03: sebelumnya
    3 ekspor (excel_rekap/pdf_rekap/pdf_dosen) selalu pakai nama akun tanpa
    gelar meski profil SIMDA sudah tersedia di baris yang sama."""

    def test_dengan_profil_pakai_nama_gelar_dan_fakultas_prodi_simda(self):
        dosen = MagicMock(kode_fakultas="FT", kode_prodi="TI")
        dosen.get_full_name.return_value = "Contoh Dosen"
        profil = MagicMock(
            nama_lengkap_gelar="Dr. Contoh Dosen, M.Kom",
            kode_fakultas="FEB", kode_prodi="MAN",
        )

        nama, fakultas, prodi = _identitas_tampilan_dosen(dosen, profil)

        self.assertEqual(nama, "Dr. Contoh Dosen, M.Kom")
        self.assertEqual(fakultas, "FEB")
        self.assertEqual(prodi, "MAN")

    def test_tanpa_profil_fallback_ke_data_akun(self):
        dosen = MagicMock(kode_fakultas="FT", kode_prodi="TI")
        dosen.get_full_name.return_value = "Contoh Dosen"

        nama, fakultas, prodi = _identitas_tampilan_dosen(dosen, None)

        self.assertEqual(nama, "Contoh Dosen")
        self.assertEqual(fakultas, "FT")
        self.assertEqual(prodi, "TI")

    def test_profil_kode_fakultas_prodi_kosong_fallback_ke_akun(self):
        dosen = MagicMock(kode_fakultas="FT", kode_prodi="TI")
        profil = MagicMock(nama_lengkap_gelar="Dr. X", kode_fakultas="", kode_prodi="")

        _, fakultas, prodi = _identitas_tampilan_dosen(dosen, profil)

        self.assertEqual(fakultas, "FT")
        self.assertEqual(prodi, "TI")


class ExportRekapNamaSimdaTest(TestCase):
    """Ekspor Excel Rekap Detail & PDF Rekap harus pakai nama+gelar dan
    fakultas/prodi dari SIMDA (profil), bukan nama akun polos tanpa gelar --
    bug fix per 2026-08-03."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin_rekap_export", password="testpass123", role="admin",
        )
        self.dosen = User.objects.create_user(
            username="dosen_rekap_export", password="testpass123", role="dosen",
            nidn="8800000099", first_name="Contoh", last_name="Dosen",
            kode_fakultas="FT", kode_prodi="TI",
        )
        self.client = Client()
        self.client.login(username="admin_rekap_export", password="testpass123")

    def _mock_profil(self):
        profil = MagicMock(
            nama_lengkap_gelar="Dr. Contoh Dosen, M.Kom",
            kode_fakultas="FEB", kode_prodi="MAN",
            persentase_kelengkapan=50, jabatan_fungsional_nama="Lektor",
            pendidikan_terakhir="S2", status_kepegawaian_nama="Aktif",
        )
        profil.riwayat_bkd.all.return_value.count.return_value = 0
        return profil

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_excel_rekap_pakai_nama_gelar_dan_fakultas_simda(self, mock_profil_fn):
        mock_profil_fn.return_value = self._mock_profil()

        response = self.client.get(reverse("laporan:excel_rekap"))

        self.assertEqual(response.status_code, 200)
        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=6, column=2).value, "Dr. Contoh Dosen, M.Kom")
        self.assertEqual(ws.cell(row=6, column=4).value, "FEB")
        self.assertEqual(ws.cell(row=6, column=5).value, "MAN")

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_pdf_rekap_tidak_error_dengan_profil_simda(self, mock_profil_fn):
        mock_profil_fn.return_value = self._mock_profil()

        response = self.client.get(reverse("laporan:pdf_rekap"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_pdf_dosen_pakai_bidang_keahlian_simda_bukan_hardcode(self, mock_profil_fn):
        profil = self._mock_profil()
        profil.bidang_keahlian_nama = "Rekayasa Perangkat Lunak"
        profil.riwayat_jabfung.all.return_value.order_by.return_value = []
        profil.riwayat_pendidikan.all.return_value.order_by.return_value = []
        mock_profil_fn.return_value = profil

        response = self.client.get(
            reverse("laporan:pdf_dosen", args=[self.dosen.id])
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_tanpa_profil_tetap_fallback_nama_akun(self, mock_profil_fn):
        mock_profil_fn.return_value = None

        response = self.client.get(reverse("laporan:excel_rekap"))

        wb = openpyxl.load_workbook(io.BytesIO(response.content))
        ws = wb.active
        self.assertEqual(ws.cell(row=6, column=2).value, "Contoh Dosen")
        self.assertEqual(ws.cell(row=6, column=4).value, "FT")


class ExportExcelStatistikProfilBucketingTest(TestCase):
    """export_excel_statistik_profil::get_profil_counts -- populasi per
    Fakultas/Prodi harus dibucket dari DataDosen.kode_fakultas/kode_prodi
    (SIMDA), bukan accounts.User -- bug fix per 2026-08-03 (sebelumnya
    dosen yang kode_fakultas User-nya belum sinkron dengan SIMDA bisa
    kehitung di fakultas yang salah)."""

    def setUp(self):
        self.admin = User.objects.create_user(
            username="admin_statistik_profil", password="testpass123", role="admin",
        )
        self.client = Client()
        self.client.login(username="admin_statistik_profil", password="testpass123")
        self.fakultas = Fakultas.objects.create(kode_fakultas="FT", nama_fakultas="Fakultas Teknik")
        Prodi.objects.create(kode_prodi="TI", nama_prodi="Teknik Informatika", fakultas=self.fakultas)

    @patch("laporan.views.JabatanFungsionalPublik")
    @patch("laporan.views.StatusKepegawaianPublik")
    @patch("laporan.views.DataDosen")
    def test_bucketing_query_datadosen_pakai_kode_fakultas_prodi(
        self, mock_datadosen, mock_status_ref, mock_jabfung_ref
    ):
        mock_status_ref.objects.using.return_value.all.return_value = []
        mock_jabfung_ref.objects.using.return_value.all.return_value = []
        mock_datadosen.objects.using.return_value.filter.return_value.count.return_value = 3
        mock_datadosen.objects.using.return_value.filter.return_value.values_list.return_value = []

        response = self.client.get(reverse("laporan:excel_statistik_profil"))

        self.assertEqual(response.status_code, 200)
        mock_datadosen.objects.using.assert_any_call("simda")
        filter_kwargs_used = [
            call.kwargs for call in mock_datadosen.objects.using.return_value.filter.call_args_list
        ]
        self.assertTrue(any(kw == {"kode_fakultas": "FT"} for kw in filter_kwargs_used))
        self.assertTrue(any(kw == {"kode_prodi": "TI"} for kw in filter_kwargs_used))


class StatistikExportAksesTest(TestCase):
    """Bug fix per 2026-08-03: export_excel_statistik_kinerja dan
    export_excel_statistik_profil sebelumnya TIDAK punya pengecekan role
    sama sekali -- siapa pun yang login (termasuk dosen/tendik) bisa
    mengunduh statistik agregat SELURUH institusi. Sekarang dibatasi ke
    admin/rektorat/biro (role yang memang berhak lihat semua fakultas/
    prodi tanpa batas cakupan)."""

    def setUp(self):
        self.dosen = User.objects.create_user(
            username="dosen_statistik", password="testpass123", role="dosen",
        )
        self.dekan = User.objects.create_user(
            username="dekan_statistik", password="testpass123", role="dekan", kode_fakultas="FT",
        )
        self.rektorat = User.objects.create_user(
            username="rektorat_statistik", password="testpass123", role="rektorat",
        )
        self.admin = User.objects.create_user(
            username="admin_statistik", password="testpass123", role="admin",
        )
        self.client = Client()

    def test_dosen_ditolak_statistik_kinerja(self):
        self.client.login(username="dosen_statistik", password="testpass123")
        response = self.client.get(reverse("laporan:excel_statistik_kinerja"))
        self.assertEqual(response.status_code, 403)

    def test_dekan_ditolak_statistik_kinerja(self):
        self.client.login(username="dekan_statistik", password="testpass123")
        response = self.client.get(reverse("laporan:excel_statistik_kinerja"))
        self.assertEqual(response.status_code, 403)

    def test_dosen_ditolak_statistik_profil(self):
        self.client.login(username="dosen_statistik", password="testpass123")
        response = self.client.get(reverse("laporan:excel_statistik_profil"))
        self.assertEqual(response.status_code, 403)

    def test_rektorat_diterima_statistik_kinerja(self):
        self.client.login(username="rektorat_statistik", password="testpass123")
        response = self.client.get(reverse("laporan:excel_statistik_kinerja"))
        self.assertEqual(response.status_code, 200)

    def test_admin_diterima_statistik_profil(self):
        self.client.login(username="admin_statistik", password="testpass123")
        response = self.client.get(reverse("laporan:excel_statistik_profil"))
        self.assertEqual(response.status_code, 200)


class ExportPdfDosenAksesTest(TestCase):
    """Bug fix per 2026-08-03: export_pdf_dosen sebelumnya cuma menolak
    role 'dosen' yang melihat dosen lain -- dekan/kaprodi bisa mengambil
    PDF dosen di luar fakultas/prodinya, dan tendik tidak dicek sama
    sekali. Sekarang dibatasi lewat get_dosen_queryset (cakupan yang
    sama dengan Rekap Data), kecuali melihat data diri sendiri."""

    def setUp(self):
        self.dosen_ft = User.objects.create_user(
            username="dosen_ft_pdfdosen", password="testpass123", role="dosen",
            kode_fakultas="FT", kode_prodi="TI",
        )
        self.dekan_ft = User.objects.create_user(
            username="dekan_ft_pdfdosen", password="testpass123", role="dekan",
            kode_fakultas="FT",
        )
        self.dekan_feb = User.objects.create_user(
            username="dekan_feb_pdfdosen", password="testpass123", role="dekan",
            kode_fakultas="FEB",
        )
        self.tendik = User.objects.create_user(
            username="tendik_pdfdosen", password="testpass123", role="tendik",
        )
        self.client = Client()

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_dekan_beda_fakultas_ditolak(self, mock_profil_fn):
        mock_profil_fn.return_value = None
        self.client.login(username="dekan_feb_pdfdosen", password="testpass123")
        response = self.client.get(reverse("laporan:pdf_dosen", args=[self.dosen_ft.id]))
        self.assertEqual(response.status_code, 403)

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_tendik_ditolak(self, mock_profil_fn):
        mock_profil_fn.return_value = None
        self.client.login(username="tendik_pdfdosen", password="testpass123")
        response = self.client.get(reverse("laporan:pdf_dosen", args=[self.dosen_ft.id]))
        self.assertEqual(response.status_code, 403)

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_dekan_fakultas_sama_diterima(self, mock_profil_fn):
        mock_profil_fn.return_value = None
        self.client.login(username="dekan_ft_pdfdosen", password="testpass123")
        response = self.client.get(reverse("laporan:pdf_dosen", args=[self.dosen_ft.id]))
        self.assertEqual(response.status_code, 200)

    @patch("laporan.views.get_simda_dosen_or_none")
    def test_dosen_lihat_diri_sendiri_diterima(self, mock_profil_fn):
        mock_profil_fn.return_value = None
        self.client.login(username="dosen_ft_pdfdosen", password="testpass123")
        response = self.client.get(reverse("laporan:pdf_dosen", args=[self.dosen_ft.id]))
        self.assertEqual(response.status_code, 200)
