from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from accounts.models import User
from master.models import Fakultas, Prodi, TahunAkademik, Pengaturan
from profil.models import ProfilDosen
from kinerja.models import Penelitian, Publikasi, PKM, HKI, BKD
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT
import io


@login_required
def index(request):
    tahun_list = TahunAkademik.objects.filter(status='aktif').order_by('-urutan')
    fakultas_list = Fakultas.objects.filter(status='aktif').order_by('kode_fakultas')
    prodi_list = Prodi.objects.filter(status='aktif').order_by('kode_prodi')
    try:
        pengaturan = Pengaturan.objects.first()
    except:
        pengaturan = None
    context = {
        'tahun_list': tahun_list,
        'fakultas_list': fakultas_list,
        'prodi_list': prodi_list,
        'pengaturan': pengaturan,
    }
    return render(request, 'laporan/index.html', context)

# ============================================================
# HELPER FUNCTIONS
# ============================================================

def get_dosen_queryset(user, filter_prodi='', filter_fakultas='', filter_status=''):
    """Ambil queryset dosen berdasarkan role dan filter"""
    if user.role in ['admin', 'rektorat', 'biro']:
        qs = User.objects.filter(role='dosen')
    elif user.role in ['dekan', 'wadek']:
        qs = User.objects.filter(role='dosen', kode_fakultas=user.kode_fakultas)
    elif user.role in ['kaprodi', 'sekprodi', 'operator']:
        qs = User.objects.filter(role='dosen', kode_prodi=user.kode_prodi)
    else:
        qs = User.objects.none()

    if filter_prodi:
        qs = qs.filter(kode_prodi=filter_prodi)
    if filter_fakultas:
        qs = qs.filter(kode_fakultas=filter_fakultas)
    if filter_status:
        qs = qs.filter(status_kepegawaian=filter_status)

    return qs.order_by('kode_fakultas', 'kode_prodi', 'first_name')


def get_tahun_range(request):
    """Ambil list tahun akademik berdasarkan filter yang dipilih"""
    tahun = request.GET.get('tahun', '')
    tahun_awal = request.GET.get('tahun_awal', '')
    tahun_akhir = request.GET.get('tahun_akhir', '')

    if tahun:
        return [tahun], tahun, tahun
    elif tahun_awal and tahun_akhir:
        from master.models import TahunAkademik
        tahun_list = TahunAkademik.objects.filter(
            tahun_akademik__gte=tahun_awal,
            tahun_akademik__lte=tahun_akhir,
            status='aktif'
        ).values_list('tahun_akademik', flat=True)
        label = f'{tahun_awal} s/d {tahun_akhir}'
        return list(tahun_list), '', label
    return [], '', 'Semua Periode'




# ============================================================
# EXPORT EXCEL
# ============================================================
@login_required

def export_excel_rekap(request):
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    
    filter_semester = request.GET.get('semester', '')
    filter_prodi = request.GET.get('prodi', '')
    filter_fakultas = request.GET.get('fakultas', '')
    filter_status = request.GET.get('status_kepegawaian', '')

    dosen_qs = get_dosen_queryset(request.user, filter_prodi, filter_fakultas, filter_status)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Kinerja Dosen"

    # Style
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Judul
    ws.merge_cells('A1:N1')
    ws['A1'] = 'REKAP DATA KINERJA DOSEN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.merge_cells('A2:N2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    filter_info = []
    if filter_tahun:
        filter_info.append(f'Tahun: {filter_tahun}')
    if filter_semester:
        filter_info.append(f'Semester: {filter_semester}')
    if filter_prodi:
        filter_info.append(f'Prodi: {filter_prodi}')
    if filter_info:
        ws.merge_cells('A3:N3')
        ws['A3'] = ' | '.join(filter_info)
        ws['A3'].alignment = center

    # Header tabel
    headers = [
        'No', 'Nama Dosen', 'NIDN', 'Fakultas', 'Prodi',
        'Jabfung', 'Pend.', 'Status Kpg.',
        'BKD', 'Penelitian', 'Publikasi', 'PKM', 'HKI', 'Profil %'
    ]
    row_header = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_header, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    ws.row_dimensions[row_header].height = 30

    # Data
    for idx, dosen in enumerate(dosen_qs, 1):
        penelitian_qs = dosen.penelitian_set.all()
        publikasi_qs = dosen.publikasi_set.all()
        pkm_qs = dosen.pkm_set.all()
        hki_qs = dosen.hki_set.all()
        bkd_qs = dosen.bkd_set.all()

        if tahun_range:
            penelitian_qs = penelitian_qs.filter(tahun_akademik__in=tahun_range)
            publikasi_qs = publikasi_qs.filter(tahun_akademik__in=tahun_range)
            pkm_qs = pkm_qs.filter(tahun_akademik__in=tahun_range)
            hki_qs = hki_qs.filter(tahun_akademik__in=tahun_range)
            bkd_qs = bkd_qs.filter(tahun_akademik__in=tahun_range)
          
            

        if filter_semester:
            penelitian_qs = penelitian_qs.filter(semester=filter_semester)
            publikasi_qs = publikasi_qs.filter(semester=filter_semester)
            pkm_qs = pkm_qs.filter(semester=filter_semester)
            hki_qs = hki_qs.filter(semester=filter_semester)
            bkd_qs = bkd_qs.filter(semester=filter_semester)

        try:
            profil = dosen.profil
            kelengkapan = profil.persentase_kelengkapan
            jabfung = profil.jabfung_aktif or '-'
            pendidikan = profil.pendidikan_terakhir or '-'
        except:
            kelengkapan = 0
            jabfung = '-'
            pendidikan = '-'

        row_data = [
            idx,
            dosen.get_full_name() or dosen.username,
            dosen.nidn or '-',
            dosen.kode_fakultas or '-',
            dosen.kode_prodi or '-',
            jabfung,
            pendidikan,
            dosen.status_kepegawaian or 'Aktif',
            bkd_qs.count(),
            penelitian_qs.count(),
            publikasi_qs.count(),
            pkm_qs.count(),
            hki_qs.count(),
            f'{kelengkapan}%',
        ]

        row_num = row_header + idx
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin
            cell.alignment = center if col != 2 else left
            if col == 9 and isinstance(value, int) and value == 0:
                cell.fill = PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid')
            elif col in [9, 10, 11, 12, 13] and isinstance(value, int) and value > 0:
                cell.fill = PatternFill(start_color='E0F0E0', end_color='E0F0E0', fill_type='solid')

    # Lebar kolom
    col_widths = [5, 30, 15, 10, 8, 20, 8, 15, 8, 12, 12, 8, 8, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Rekap_Kinerja_Dosen_{filter_tahun or "Semua"}_{filter_semester or "Semua"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_excel_penelitian(request):
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    filter_semester = request.GET.get('semester', '')
    filter_prodi = request.GET.get('prodi', '')
    filter_fakultas = request.GET.get('fakultas', '')

    dosen_qs = get_dosen_queryset(request.user, filter_prodi, filter_fakultas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Penelitian"

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:K1')
    ws['A1'] = 'DATA PENELITIAN DOSEN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center
    ws.merge_cells('A2:K2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    headers = ['No', 'Nama Dosen', 'NIDN', 'Prodi', 'Judul', 'Semester',
               'Tahun Akademik', 'L/N/I', 'Sumber Dana', 'Pendanaan (Juta)', 'Jml Mahasiswa']
    row_h = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    row_num = row_h + 1
    no = 1
    for dosen in dosen_qs:
        penelitian_qs = dosen.penelitian_set.all()
        if tahun_range:
            penelitian_qs = penelitian_qs.filter(tahun_akademik__in=tahun_range)
        if filter_semester:
            penelitian_qs = penelitian_qs.filter(semester=filter_semester)

        for p in penelitian_qs:
            row_data = [
                no, dosen.get_full_name() or dosen.username,
                dosen.nidn or '-', dosen.kode_prodi or '-',
                p.judul, p.semester or '-', p.tahun_akademik,
                p.get_ln_i_display() if p.ln_i else '-',
                p.sumber or '-', float(p.pendanaan), p.jml_mahasiswa
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin
                cell.alignment = left if col in [5] else center
            row_num += 1
            no += 1

    col_widths = [5, 25, 15, 8, 50, 10, 15, 8, 20, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Data_Penelitian_{filter_tahun or "Semua"}_{filter_semester or "Semua"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_excel_publikasi(request):
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    
    filter_semester = request.GET.get('semester', '')
    filter_prodi = request.GET.get('prodi', '')
    filter_fakultas = request.GET.get('fakultas', '')

    dosen_qs = get_dosen_queryset(request.user, filter_prodi, filter_fakultas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data Publikasi"

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:K1')
    ws['A1'] = 'DATA PUBLIKASI DOSEN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center
    ws.merge_cells('A2:K2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    headers = ['No', 'Nama Dosen', 'NIDN', 'Prodi', 'Judul',
               'Jenis', 'Nama Jurnal', 'Volume', 'Nomor',
               'Tahun Terbit', 'Semester', 'Tahun Akademik']
    row_h = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin

    row_num = row_h + 1
    no = 1
    for dosen in dosen_qs:
        publikasi_qs = dosen.publikasi_set.all()
        if tahun_range:
            publikasi_qs = publikasi_qs.filter(tahun_akademik__in=tahun_range)
        if filter_semester:
            publikasi_qs = publikasi_qs.filter(semester=filter_semester)

        for p in publikasi_qs:
            row_data = [
                no, dosen.get_full_name() or dosen.username,
                dosen.nidn or '-', dosen.kode_prodi or '-',
                p.judul, p.jenis_publikasi,
                p.nama_jurnal or '-', p.volume or '-',
                p.nomor or '-', p.tahun_terbit or '-',
                p.semester or '-', p.tahun_akademik
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin
                cell.alignment = left if col == 5 else center
            row_num += 1
            no += 1

    col_widths = [5, 25, 15, 8, 50, 8, 30, 8, 8, 12, 10, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Data_Publikasi_{filter_tahun or "Semua"}_{filter_semester or "Semua"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ============================================================
# EXPORT PDF
# ============================================================
@login_required
def export_pdf_rekap(request):
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    filter_semester = request.GET.get('semester', '')
    filter_prodi = request.GET.get('prodi', '')
    filter_fakultas = request.GET.get('fakultas', '')
    filter_status = request.GET.get('status_kepegawaian', '')

    dosen_qs = get_dosen_queryset(request.user, filter_prodi, filter_fakultas, filter_status)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        rightMargin=1*cm, leftMargin=1*cm,
        topMargin=1.5*cm, bottomMargin=1*cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', parent=styles['Heading1'],
                                  fontSize=14, alignment=TA_CENTER, spaceAfter=4)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                fontSize=10, alignment=TA_CENTER, spaceAfter=2)
    small_style = ParagraphStyle('small', parent=styles['Normal'],
                                  fontSize=8, alignment=TA_LEFT)

    elements = []
    elements.append(Paragraph('REKAP DATA KINERJA DOSEN', title_style))
    elements.append(Paragraph('Universitas Ichsan Gorontalo', sub_style))

    # Info periode untuk header PDF
    filter_info = []
    if periode_label:
        filter_info.append(f'Periode: {periode_label}')
    elif filter_tahun:
        filter_info.append(f'Tahun: {filter_tahun}')
    if filter_semester:
        filter_info.append(f'Semester: {filter_semester}')
    if filter_prodi:
        filter_info.append(f'Prodi: {filter_prodi}')
    if filter_info:
        elements.append(Paragraph(' | '.join(filter_info), sub_style))

   
    elements.append(Spacer(1, 0.3*cm))

    header_color = colors.HexColor('#1e3a5f')
    alt_color = colors.HexColor('#f0f4f8')

    table_data = [['No', 'Nama Dosen', 'NIDN', 'Prodi', 'Jabfung',
                   'Pend.', 'Status', 'BKD', 'Penelitian', 'Publikasi', 'PKM', 'HKI', 'Profil%']]

    for idx, dosen in enumerate(dosen_qs, 1):
        penelitian_qs = dosen.penelitian_set.all()
        publikasi_qs = dosen.publikasi_set.all()
        pkm_qs = dosen.pkm_set.all()
        hki_qs = dosen.hki_set.all()
        bkd_qs = dosen.bkd_set.all()

        if tahun_range:
            penelitian_qs = penelitian_qs.filter(tahun_akademik__in=tahun_range)
            publikasi_qs = publikasi_qs.filter(tahun_akademik__in=tahun_range)
            pkm_qs = pkm_qs.filter(tahun_akademik__in=tahun_range)
            hki_qs = hki_qs.filter(tahun_akademik__in=tahun_range)
            bkd_qs = bkd_qs.filter(tahun_akademik__in=tahun_range)
         

        if filter_semester:
            penelitian_qs = penelitian_qs.filter(semester=filter_semester)
            publikasi_qs = publikasi_qs.filter(semester=filter_semester)
            pkm_qs = pkm_qs.filter(semester=filter_semester)
            hki_qs = hki_qs.filter(semester=filter_semester)
            bkd_qs = bkd_qs.filter(semester=filter_semester)

        try:
            profil = dosen.profil
            kelengkapan = profil.persentase_kelengkapan
            jabfung = profil.jabfung_aktif or '-'
            pendidikan = profil.pendidikan_terakhir or '-'
        except:
            kelengkapan = 0
            jabfung = '-'
            pendidikan = '-'

        table_data.append([
            str(idx),
            Paragraph(dosen.get_full_name() or dosen.username, small_style),
            dosen.nidn or '-',
            dosen.kode_prodi or '-',
            jabfung,
            pendidikan,
            dosen.status_kepegawaian or 'Aktif',
            str(bkd_qs.count()),
            str(penelitian_qs.count()),
            str(publikasi_qs.count()),
            str(pkm_qs.count()),
            str(hki_qs.count()),
            f'{kelengkapan}%',
        ])

    col_widths_pdf = [1*cm, 5*cm, 2.5*cm, 1.8*cm, 3*cm,
                      1.5*cm, 2.5*cm, 1.2*cm, 1.8*cm, 1.8*cm, 1.2*cm, 1.2*cm, 1.5*cm]

    table = Table(table_data, colWidths=col_widths_pdf, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), header_color),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, alt_color]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))

    elements.append(table)
    doc.build(elements)

    buffer.seek(0)
    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f'Rekap_Kinerja_Dosen_{filter_tahun or "Semua"}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def export_excel_pkm(request):
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    filter_semester = request.GET.get('semester', '')
    filter_prodi = request.GET.get('prodi', '')
    filter_fakultas = request.GET.get('fakultas', '')

    dosen_qs = get_dosen_queryset(request.user, filter_prodi, filter_fakultas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data PKM"

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:K1')
    ws['A1'] = 'DATA PENGABDIAN KEPADA MASYARAKAT (PKM) DOSEN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.merge_cells('A2:K2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    filter_info = []
    if filter_tahun:
        filter_info.append(f'Tahun: {filter_tahun}')
    if filter_semester:
        filter_info.append(f'Semester: {filter_semester}')
    if filter_info:
        ws.merge_cells('A3:K3')
        ws['A3'] = ' | '.join(filter_info)
        ws['A3'].alignment = center

    headers = [
        'No', 'Nama Dosen', 'NIDN', 'Prodi', 'Judul PKM',
        'Semester', 'Tahun Akademik', 'L/N/I',
        'Sumber Dana', 'Pendanaan (Juta)', 'Jml Mahasiswa'
    ]
    row_h = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_h].height = 30

    row_num = row_h + 1
    no = 1
    for dosen in dosen_qs:
        pkm_qs = dosen.pkm_set.all()
        if tahun_range:
            pkm_qs = pkm_qs.filter(tahun_akademik__in=tahun_range)
        if filter_semester:
            pkm_qs = pkm_qs.filter(semester=filter_semester)

        for p in pkm_qs:
            row_data = [
                no,
                dosen.get_full_name() or dosen.username,
                dosen.nidn or '-',
                dosen.kode_prodi or '-',
                p.judul,
                p.semester or '-',
                p.tahun_akademik,
                p.get_ln_i_display() if p.ln_i else '-',
                p.sumber or '-',
                float(p.pendanaan),
                p.jml_mahasiswa
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin
                cell.alignment = left if col == 5 else center
            row_num += 1
            no += 1

    if no == 1:
        ws.merge_cells(f'A{row_num}:K{row_num}')
        ws[f'A{row_num}'] = 'Tidak ada data'
        ws[f'A{row_num}'].alignment = center

    col_widths = [5, 25, 15, 8, 50, 10, 15, 8, 20, 15, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Data_PKM_{filter_tahun or "Semua"}_{filter_semester or "Semua"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_excel_hki(request):
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    filter_semester = request.GET.get('semester', '')
    filter_prodi = request.GET.get('prodi', '')
    filter_fakultas = request.GET.get('fakultas', '')

    dosen_qs = get_dosen_queryset(request.user, filter_prodi, filter_fakultas)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data HKI"

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    ws.merge_cells('A1:J1')
    ws['A1'] = 'DATA HAK KEKAYAAN INTELEKTUAL (HKI) DOSEN'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = center

    ws.merge_cells('A2:J2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    filter_info = []
    if filter_tahun:
        filter_info.append(f'Tahun: {filter_tahun}')
    if filter_semester:
        filter_info.append(f'Semester: {filter_semester}')
    if filter_info:
        ws.merge_cells('A3:J3')
        ws['A3'] = ' | '.join(filter_info)
        ws['A3'].alignment = center

    headers = [
        'No', 'Nama Dosen', 'NIDN', 'Prodi',
        'Judul HKI', 'Jenis HKI', 'No. HKI',
        'Tahun Perolehan', 'Semester', 'Tahun Akademik'
    ]
    row_h = 5
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_h, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_h].height = 30

    row_num = row_h + 1
    no = 1
    for dosen in dosen_qs:
        hki_qs = dosen.hki_set.all()
        if filter_tahun:
            hki_qs = hki_qs.filter(tahun_akademik=filter_tahun)
        if filter_semester:
            hki_qs = hki_qs.filter(semester=filter_semester)

        for h in hki_qs:
            row_data = [
                no,
                dosen.get_full_name() or dosen.username,
                dosen.nidn or '-',
                dosen.kode_prodi or '-',
                h.judul,
                h.jenis_hki,
                h.no_hki or '-',
                h.tahun_perolehan or '-',
                h.semester or '-',
                h.tahun_akademik,
            ]
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin
                cell.alignment = left if col == 5 else center
            row_num += 1
            no += 1

    if no == 1:
        ws.merge_cells(f'A{row_num}:J{row_num}')
        ws[f'A{row_num}'] = 'Tidak ada data'
        ws[f'A{row_num}'].alignment = center

    col_widths = [5, 25, 15, 8, 50, 20, 20, 15, 10, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Data_HKI_{filter_tahun or "Semua"}_{filter_semester or "Semua"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response

@login_required
def export_pdf_dosen(request, dosen_id):
    from django.shortcuts import get_object_or_404
    target_dosen = get_object_or_404(User, id=dosen_id)

    # Cek akses — dosen hanya bisa lihat miliknya sendiri
    if request.user.role == 'dosen' and request.user.id != target_dosen.id:
        messages.error(request, 'Anda tidak memiliki akses laporan dosen lain.')
        return redirect('kinerja:index')
        
    # Cek akses
    if request.user.role == 'dosen' and request.user != target_dosen:
        from django.http import HttpResponseForbidden
        return HttpResponseForbidden()
    tahun_range, filter_tahun, periode_label = get_tahun_range(request)
    filter_semester = request.GET.get('semester', '')

    # Ambil data
    try:
        profil = target_dosen.profil
    except:
        profil = None

    jabfung_list = target_dosen.jabfung_set.all().order_by('-tgl_sk')
    pendidikan_list = target_dosen.pendidikan_set.all().order_by('-tahun_lulus')
    sertifikat_list = target_dosen.sertifikat_set.all().order_by('-tahun_terbit')

    penelitian_qs = target_dosen.penelitian_set.all()
    publikasi_qs = target_dosen.publikasi_set.all()
    pkm_qs = target_dosen.pkm_set.all()
    hki_qs = target_dosen.hki_set.all()
    bkd_qs = target_dosen.bkd_set.all()

    if tahun_range:
        penelitian_qs = penelitian_qs.filter(tahun_akademik__in=tahun_range)
        publikasi_qs = publikasi_qs.filter(tahun_akademik__in=tahun_range)
        pkm_qs = pkm_qs.filter(tahun_akademik__in=tahun_range)
        hki_qs = hki_qs.filter(tahun_akademik__in=tahun_range)
        bkd_qs = bkd_qs.filter(tahun_akademik__in=tahun_range)

    if filter_semester:
        penelitian_qs = penelitian_qs.filter(semester=filter_semester)
        publikasi_qs = publikasi_qs.filter(semester=filter_semester)
        pkm_qs = pkm_qs.filter(semester=filter_semester)
        hki_qs = hki_qs.filter(semester=filter_semester)
        bkd_qs = bkd_qs.filter(semester=filter_semester)

    # Build PDF
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm
    )

    styles = getSampleStyleSheet()
    header_color = colors.HexColor('#1e3a5f')
    alt_color = colors.HexColor('#f0f4f8')

    title_style = ParagraphStyle('title', parent=styles['Normal'],
        fontSize=14, fontName='Helvetica-Bold',
        alignment=TA_CENTER, spaceAfter=4, textColor=colors.white)
    sub_style = ParagraphStyle('sub', parent=styles['Normal'],
        fontSize=10, alignment=TA_CENTER,
        spaceAfter=2, textColor=colors.white)
    section_style = ParagraphStyle('section', parent=styles['Normal'],
        fontSize=11, fontName='Helvetica-Bold',
        textColor=colors.white, spaceAfter=4)
    normal_style = ParagraphStyle('normal', parent=styles['Normal'],
        fontSize=9, spaceAfter=2)
    small_style = ParagraphStyle('small', parent=styles['Normal'],
        fontSize=8)

    elements = []

    # Header
    header_data = [[
        Paragraph(
            f'<b>LAPORAN KINERJA DOSEN</b><br/>'
            f'Universitas Ichsan Gorontalo<br/>'
            f'<font size="9">'
            f'{"Periode: " + filter_semester + " " + filter_tahun if filter_tahun else "Semua Periode"}'
            f'</font>',
            ParagraphStyle('hdr', fontSize=13, fontName='Helvetica-Bold',
                          alignment=TA_CENTER, textColor=colors.white)
        )
    ]]
    header_table = Table(header_data, colWidths=[17*cm])
    header_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), header_color),
        ('TOPPADDING', (0,0), (-1,-1), 12),
        ('BOTTOMPADDING', (0,0), (-1,-1), 12),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.4*cm))

    # Identitas Dosen
    nama = target_dosen.get_full_name() or target_dosen.username
    nidn = target_dosen.nidn or '-'
    prodi = target_dosen.kode_prodi or '-'
    fakultas = target_dosen.kode_fakultas or '-'
    jabfung_aktif = profil.jabfung_aktif if profil else '-'
    pendidikan = profil.pendidikan_terakhir if profil else '-'
    bidang = profil.bidang_keahlian if profil else '-'
    status_kpg = target_dosen.status_kepegawaian or 'Aktif'

    identitas_data = [
        ['DATA IDENTITAS DOSEN', '', '', ''],
        ['Nama Lengkap', f': {nama}', 'Prodi', f': {prodi}'],
        ['NIDN', f': {nidn}', 'Fakultas', f': {fakultas}'],
        ['Jabfung', f': {jabfung_aktif}', 'Pend. Terakhir', f': {pendidikan}'],
        ['Bidang Keahlian', f': {bidang}', 'Status', f': {status_kpg}'],
    ]

    id_table = Table(identitas_data, colWidths=[3.5*cm, 5*cm, 3.5*cm, 5*cm])
    id_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('SPAN', (0,0), (-1,0)),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
    ]))
    elements.append(id_table)
    elements.append(Spacer(1, 0.4*cm))

    # Helper fungsi section header
    def section_header(title):
        data = [[Paragraph(f'<b>{title}</b>',
                 ParagraphStyle('sh', fontSize=10, fontName='Helvetica-Bold',
                               textColor=colors.white))]]
        t = Table(data, colWidths=[17*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#2d6a9f')),
            ('TOPPADDING', (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
        ]))
        return t

    # BKD
    elements.append(section_header('1. BEBAN KERJA DOSEN (BKD)'))
    elements.append(Spacer(1, 0.2*cm))
    if bkd_qs.exists():
        bkd_data = [['No', 'Semester', 'Tahun Akademik', 'Status']]
        for i, b in enumerate(bkd_qs, 1):
            status = 'Terupload' if (b.file_bkd or b.link_bkd) else 'Belum Upload'
            bkd_data.append([str(i), b.semester, b.tahun_akademik, status])
        t = Table(bkd_data, colWidths=[1*cm, 4*cm, 6*cm, 6*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph('Tidak ada data BKD.', small_style))
    elements.append(Spacer(1, 0.4*cm))

    # Penelitian
    elements.append(section_header('2. PENELITIAN'))
    elements.append(Spacer(1, 0.2*cm))
    if penelitian_qs.exists():
        pen_data = [['No', 'Judul', 'Semester', 'Tahun', 'L/N/I', 'Dana (Juta)']]
        for i, p in enumerate(penelitian_qs, 1):
            pen_data.append([
                str(i),
                Paragraph(p.judul[:80], small_style),
                p.semester or '-',
                p.tahun_akademik,
                p.get_ln_i_display() if p.ln_i else '-',
                f'Rp {p.pendanaan}'
            ])
        t = Table(pen_data, colWidths=[1*cm, 7*cm, 2*cm, 2.5*cm, 1.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph('Tidak ada data penelitian.', small_style))
    elements.append(Spacer(1, 0.4*cm))

    # Publikasi
    elements.append(section_header('3. PUBLIKASI'))
    elements.append(Spacer(1, 0.2*cm))
    if publikasi_qs.exists():
        pub_data = [['No', 'Judul', 'Jenis', 'Jurnal', 'Tahun']]
        for i, p in enumerate(publikasi_qs, 1):
            pub_data.append([
                str(i),
                Paragraph(p.judul[:80], small_style),
                p.jenis_publikasi,
                Paragraph(p.nama_jurnal[:40] if p.nama_jurnal else '-', small_style),
                str(p.tahun_terbit or '-')
            ])
        t = Table(pub_data, colWidths=[1*cm, 7*cm, 2*cm, 4.5*cm, 2.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph('Tidak ada data publikasi.', small_style))
    elements.append(Spacer(1, 0.4*cm))

    # PKM
    elements.append(section_header('4. PENGABDIAN KEPADA MASYARAKAT (PKM)'))
    elements.append(Spacer(1, 0.2*cm))
    if pkm_qs.exists():
        pkm_data = [['No', 'Judul', 'Semester', 'Tahun', 'L/N/I', 'Dana (Juta)']]
        for i, p in enumerate(pkm_qs, 1):
            pkm_data.append([
                str(i),
                Paragraph(p.judul[:80], small_style),
                p.semester or '-',
                p.tahun_akademik,
                p.get_ln_i_display() if p.ln_i else '-',
                f'Rp {p.pendanaan}'
            ])
        t = Table(pkm_data, colWidths=[1*cm, 7*cm, 2*cm, 2.5*cm, 1.5*cm, 3*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph('Tidak ada data PKM.', small_style))
    elements.append(Spacer(1, 0.4*cm))

    # HKI
    elements.append(section_header('5. HAK KEKAYAAN INTELEKTUAL (HKI)'))
    elements.append(Spacer(1, 0.2*cm))
    if hki_qs.exists():
        hki_data = [['No', 'Judul', 'Jenis', 'No. HKI', 'Tahun']]
        for i, h in enumerate(hki_qs, 1):
            hki_data.append([
                str(i),
                Paragraph(h.judul[:80], small_style),
                h.jenis_hki,
                h.no_hki or '-',
                str(h.tahun_perolehan or '-')
            ])
        t = Table(hki_data, colWidths=[1*cm, 7*cm, 3*cm, 3.5*cm, 2.5*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), header_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('ALIGN', (0,0), (0,-1), 'CENTER'),
            ('ALIGN', (0,0), (-1,0), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
            ('TOPPADDING', (0,0), (-1,-1), 3),
            ('BOTTOMPADDING', (0,0), (-1,-1), 3),
        ]))
        elements.append(t)
    else:
        elements.append(Paragraph('Tidak ada data HKI.', small_style))
    elements.append(Spacer(1, 0.4*cm))

    # Ringkasan
    elements.append(section_header('RINGKASAN KINERJA'))
    elements.append(Spacer(1, 0.2*cm))
    summary_data = [
        ['Komponen', 'Jumlah'],
        ['BKD Terupload', str(bkd_qs.count())],
        ['Penelitian', str(penelitian_qs.count())],
        ['Publikasi', str(publikasi_qs.count())],
        ['PKM', str(pkm_qs.count())],
        ['HKI', str(hki_qs.count())],
    ]
    t = Table(summary_data, colWidths=[10*cm, 7*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), header_color),
        ('TEXTCOLOR', (0,0), (-1,0), colors.white),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, alt_color]),
        ('TOPPADDING', (0,0), (-1,-1), 4),
        ('BOTTOMPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(t)
    elements.append(Spacer(1, 0.6*cm))

    # Tanda tangan
    from datetime import date
    ttd_data = [[
        Paragraph('Mengetahui,<br/>Kepala Program Studi', small_style),
        Paragraph('', small_style),
        Paragraph(f'Gorontalo, {date.today().strftime("%d %B %Y")}<br/>Yang Bersangkutan', small_style),
    ]]
    ttd_table = Table(ttd_data, colWidths=[5.5*cm, 6*cm, 5.5*cm])
    ttd_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 4),
    ]))
    elements.append(ttd_table)

    # Spacer untuk tanda tangan
    elements.append(Spacer(1, 1.5*cm))
    ttd_nama = [[
        Paragraph('(________________________)', small_style),
        Paragraph('', small_style),
        Paragraph(f'({nama})', small_style),
    ]]
    ttd_nama_table = Table(ttd_nama, colWidths=[5.5*cm, 6*cm, 5.5*cm])
    ttd_nama_table.setStyle(TableStyle([
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(ttd_nama_table)

    doc.build(elements)
    buffer.seek(0)

    response = HttpResponse(buffer, content_type='application/pdf')
    filename = f'Laporan_Kinerja_{nama.replace(" ", "_")}_{filter_tahun or "Semua"}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

@login_required
def export_excel_statistik_kinerja(request):
    filter_tahun = request.GET.get('tahun', '')
    filter_tahun_awal = request.GET.get('tahun_awal', '')
    filter_tahun_akhir = request.GET.get('tahun_akhir', '')
    filter_semester = request.GET.get('semester', '')

    # Hitung tahun range
    if filter_tahun:
        tahun_range = [filter_tahun]
    elif filter_tahun_awal and filter_tahun_akhir:
        from master.models import TahunAkademik
        tahun_range = list(TahunAkademik.objects.filter(
            tahun_akademik__gte=filter_tahun_awal,
            tahun_akademik__lte=filter_tahun_akhir,
            status='aktif'
        ).values_list('tahun_akademik', flat=True))
    else:
        tahun_range = []

    from master.models import Fakultas, Prodi
    from kinerja.models import Penelitian, Publikasi, PKM, HKI, BKD

    fakultas_list = Fakultas.objects.filter(status='aktif').order_by('kode_fakultas')
    prodi_list = Prodi.objects.filter(status='aktif').order_by(
        'fakultas__kode_fakultas', 'kode_prodi'
    )

    def get_counts(filter_kwargs):
        p_qs = Penelitian.objects.filter(**filter_kwargs)
        pub_qs = Publikasi.objects.filter(**filter_kwargs)
        pkm_qs = PKM.objects.filter(**filter_kwargs)
        hki_qs = HKI.objects.filter(**filter_kwargs)

        if tahun_range:
            p_qs = p_qs.filter(tahun_akademik__in=tahun_range)
            pub_qs = pub_qs.filter(tahun_akademik__in=tahun_range)
            pkm_qs = pkm_qs.filter(tahun_akademik__in=tahun_range)
            hki_qs = hki_qs.filter(tahun_akademik__in=tahun_range)

        if filter_semester:
            p_qs = p_qs.filter(semester=filter_semester)
            pub_qs = pub_qs.filter(semester=filter_semester)
            pkm_qs = pkm_qs.filter(semester=filter_semester)
            hki_qs = hki_qs.filter(semester=filter_semester)

        # Publikasi per jenis
        pub_ib = pub_qs.filter(jenis_publikasi='IB').count()
        pub_i = pub_qs.filter(jenis_publikasi='I').count()
        pub_s1 = pub_qs.filter(jenis_publikasi='S1').count()
        pub_s2 = pub_qs.filter(jenis_publikasi='S2').count()
        pub_s3 = pub_qs.filter(jenis_publikasi='S3').count()
        pub_s4 = pub_qs.filter(jenis_publikasi='S4').count()
        pub_t = pub_qs.filter(jenis_publikasi='T').count()

        # HKI per jenis
        hki_paten = hki_qs.filter(jenis_hki='Paten').count()
        hki_ps = hki_qs.filter(jenis_hki='Paten Sederhana').count()
        hki_hc = hki_qs.filter(jenis_hki='Hak Cipta').count()
        hki_merek = hki_qs.filter(jenis_hki='Merek').count()
        hki_di = hki_qs.filter(jenis_hki='Desain Industri').count()
        hki_lain = hki_qs.filter(jenis_hki='Lainnya').count()

        return {
            'penelitian': p_qs.count(),
            'pub_total': pub_qs.count(),
            'pub_ib': pub_ib, 'pub_i': pub_i,
            'pub_s1': pub_s1, 'pub_s2': pub_s2,
            'pub_s3': pub_s3, 'pub_s4': pub_s4,
            'pub_t': pub_t,
            'pkm': pkm_qs.count(),
            'hki_total': hki_qs.count(),
            'hki_paten': hki_paten, 'hki_ps': hki_ps,
            'hki_hc': hki_hc, 'hki_merek': hki_merek,
            'hki_di': hki_di, 'hki_lain': hki_lain,
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Kinerja"

    # Style
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    subheader_fill = PatternFill(start_color='2d6a9f', end_color='2d6a9f', fill_type='solid')
    fakultas_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
    total_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Judul
    ws.merge_cells('A1:T1')
    ws['A1'] = 'REKAP STATISTIK KINERJA DOSEN'
    ws['A1'].font = Font(bold=True, size=14, color='1e3a5f')
    ws['A1'].alignment = center

    ws.merge_cells('A2:T2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    # Info periode
    periode_info = []
    if filter_tahun:
        periode_info.append(f'Tahun: {filter_tahun}')
    elif filter_tahun_awal and filter_tahun_akhir:
        periode_info.append(f'Periode: {filter_tahun_awal} s/d {filter_tahun_akhir}')
    if filter_semester:
        periode_info.append(f'Semester: {filter_semester}')
    if not periode_info:
        periode_info.append('Semua Periode')

    ws.merge_cells('A3:T3')
    ws['A3'] = ' | '.join(periode_info)
    ws['A3'].alignment = center
    ws['A3'].font = Font(italic=True, size=10)

    # Header kolom — baris 5 & 6
    headers_row1 = [
        ('A5', 'B5', 'No'),
        ('B5', 'C5', 'Fakultas / Prodi'),
        ('C5', 'D5', 'Penelitian'),
        ('D5', 'K5', 'Publikasi'),
        ('K5', 'L5', 'PKM'),
        ('L5', 'T5', 'HKI'),
    ]

    # Row 5 — header utama
    ws.merge_cells('A5:A6')
    ws['A5'] = 'No'
    ws['A5'].font = header_font
    ws['A5'].fill = header_fill
    ws['A5'].alignment = center
    ws['A5'].border = thin

    ws.merge_cells('B5:B6')
    ws['B5'] = 'Fakultas / Program Studi'
    ws['B5'].font = header_font
    ws['B5'].fill = header_fill
    ws['B5'].alignment = center
    ws['B5'].border = thin

    ws.merge_cells('C5:C6')
    ws['C5'] = 'Penelitian'
    ws['C5'].font = header_font
    ws['C5'].fill = header_fill
    ws['C5'].alignment = center
    ws['C5'].border = thin

    ws.merge_cells('D5:K5')
    ws['D5'] = 'Publikasi'
    ws['D5'].font = header_font
    ws['D5'].fill = header_fill
    ws['D5'].alignment = center
    ws['D5'].border = thin

    ws.merge_cells('L5:L6')
    ws['L5'] = 'PKM'
    ws['L5'].font = header_font
    ws['L5'].fill = header_fill
    ws['L5'].alignment = center
    ws['L5'].border = thin

    ws.merge_cells('M5:T5')
    ws['M5'] = 'HKI'
    ws['M5'].font = header_font
    ws['M5'].fill = header_fill
    ws['M5'].alignment = center
    ws['M5'].border = thin

    # Row 6 — sub header
    pub_sub = ['Total', 'Int. Bereputasi', 'Internasional',
               'Sinta 1', 'Sinta 2', 'Sinta 3', 'Sinta 4', 'Tdk Terakreditasi']
    hki_sub = ['Total', 'Paten', 'Paten Sederhana',
               'Hak Cipta', 'Merek', 'Desain Industri', 'Lainnya']

    for col, label in enumerate(pub_sub, start=4):  # D=4
        cell = ws.cell(row=6, column=col, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin

    for col, label in enumerate(hki_sub, start=13):  # M=13
        cell = ws.cell(row=6, column=col, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin

    ws.row_dimensions[5].height = 25
    ws.row_dimensions[6].height = 35

    # Data
    row_num = 7
    grand_total = {k: 0 for k in [
        'penelitian', 'pub_total', 'pub_ib', 'pub_i',
        'pub_s1', 'pub_s2', 'pub_s3', 'pub_s4', 'pub_t',
        'pkm', 'hki_total', 'hki_paten', 'hki_ps',
        'hki_hc', 'hki_merek', 'hki_di', 'hki_lain'
    ]}
    no = 1

    for fak in fakultas_list:
        # Hitung total fakultas
        fak_counts = get_counts({'kode_fakultas': fak.kode_fakultas})

        # Baris fakultas
        fak_font = Font(bold=True, size=10, color='1e3a5f')
        row_data_fak = [
            no, f'[{fak.kode_fakultas}] {fak.nama_fakultas}',
            fak_counts['penelitian'],
            fak_counts['pub_total'], fak_counts['pub_ib'],
            fak_counts['pub_i'], fak_counts['pub_s1'],
            fak_counts['pub_s2'], fak_counts['pub_s3'],
            fak_counts['pub_s4'], fak_counts['pub_t'],
            fak_counts['pkm'],
            fak_counts['hki_total'], fak_counts['hki_paten'],
            fak_counts['hki_ps'], fak_counts['hki_hc'],
            fak_counts['hki_merek'], fak_counts['hki_di'],
            fak_counts['hki_lain'],
        ]
        for col, val in enumerate(row_data_fak, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = fak_font
            cell.fill = fakultas_fill
            cell.alignment = center if col != 2 else left
            cell.border = thin
        ws.row_dimensions[row_num].height = 20
        row_num += 1
        no += 1

        # Update grand total
        for k in grand_total:
            grand_total[k] += fak_counts[k]

        # Baris per prodi dalam fakultas
        prodi_fak = prodi_list.filter(fakultas=fak)
        for prodi in prodi_fak:
            prodi_counts = get_counts({'kode_prodi': prodi.kode_prodi})
            row_data_prodi = [
                '', f'    {prodi.kode_prodi} - {prodi.nama_prodi}',
                prodi_counts['penelitian'],
                prodi_counts['pub_total'], prodi_counts['pub_ib'],
                prodi_counts['pub_i'], prodi_counts['pub_s1'],
                prodi_counts['pub_s2'], prodi_counts['pub_s3'],
                prodi_counts['pub_s4'], prodi_counts['pub_t'],
                prodi_counts['pkm'],
                prodi_counts['hki_total'], prodi_counts['hki_paten'],
                prodi_counts['hki_ps'], prodi_counts['hki_hc'],
                prodi_counts['hki_merek'], prodi_counts['hki_di'],
                prodi_counts['hki_lain'],
            ]
            for col, val in enumerate(row_data_prodi, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = Font(size=9)
                cell.alignment = center if col != 2 else left
                cell.border = thin
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # Baris Grand Total
    total_font = Font(bold=True, color='FFFFFF', size=10)
    ws.cell(row=row_num, column=1, value='').border = thin
    total_label = ws.cell(row=row_num, column=2, value='TOTAL KESELURUHAN')
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = left
    total_label.border = thin

    total_values = [
        grand_total['penelitian'],
        grand_total['pub_total'], grand_total['pub_ib'],
        grand_total['pub_i'], grand_total['pub_s1'],
        grand_total['pub_s2'], grand_total['pub_s3'],
        grand_total['pub_s4'], grand_total['pub_t'],
        grand_total['pkm'],
        grand_total['hki_total'], grand_total['hki_paten'],
        grand_total['hki_ps'], grand_total['hki_hc'],
        grand_total['hki_merek'], grand_total['hki_di'],
        grand_total['hki_lain'],
    ]
    for col, val in enumerate(total_values, start=3):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_num].height = 22

    # Lebar kolom
    col_widths = [5, 35, 12, 10, 14, 14, 10, 10, 10, 10, 16,
                  10, 10, 10, 15, 12, 10, 15, 10]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header
    ws.freeze_panes = 'C7'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'Rekap_Statistik_Kinerja_{filter_tahun or "Semua"}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def export_excel_statistik_profil(request):
    from master.models import Fakultas, Prodi

    fakultas_list = Fakultas.objects.filter(status='aktif').order_by('kode_fakultas')
    prodi_list = Prodi.objects.filter(status='aktif').order_by(
        'fakultas__kode_fakultas', 'kode_prodi'
    )

    from profil.models import ProfilDosen

    def get_profil_counts(filter_kwargs):
        dosen_qs = User.objects.filter(role='dosen', **filter_kwargs)
        total = dosen_qs.count()

        # Pendidikan
        s1 = ProfilDosen.objects.filter(
            user__in=dosen_qs, pendidikan_terakhir='S1'
        ).count()
        s2 = ProfilDosen.objects.filter(
            user__in=dosen_qs, pendidikan_terakhir='S2'
        ).count()
        s3 = ProfilDosen.objects.filter(
            user__in=dosen_qs, pendidikan_terakhir='S3'
        ).count()

        # Jabfung
        tp = ProfilDosen.objects.filter(
            user__in=dosen_qs, jabfung_aktif='Tenaga Pengajar'
        ).count()
        aa = ProfilDosen.objects.filter(
            user__in=dosen_qs, jabfung_aktif='Asisten Ahli'
        ).count()
        lek = ProfilDosen.objects.filter(
            user__in=dosen_qs, jabfung_aktif='Lektor'
        ).count()
        lk = ProfilDosen.objects.filter(
            user__in=dosen_qs, jabfung_aktif='Lektor Kepala'
        ).count()
        gb = ProfilDosen.objects.filter(
            user__in=dosen_qs, jabfung_aktif='Guru Besar'
        ).count()

        # Sertifikasi dosen
        from profil.models import Sertifikat
        serdos = Sertifikat.objects.filter(
            user__in=dosen_qs, jenis_sertifikat='Serdos'
        ).values('user').distinct().count()

        # Status kepegawaian
        aktif = dosen_qs.filter(status_kepegawaian='Aktif').count()
        tb = dosen_qs.filter(status_kepegawaian='Tugas Belajar').count()
        ls = dosen_qs.filter(status_kepegawaian='Lanjut Studi').count()
        keluar = dosen_qs.filter(status_kepegawaian='Keluar').count()
        meninggal = dosen_qs.filter(status_kepegawaian='Meninggal').count()

        return {
            'total': total,
            's1': s1, 's2': s2, 's3': s3,
            'tp': tp, 'aa': aa, 'lektor': lek,
            'lk': lk, 'gb': gb,
            'serdos': serdos,
            'aktif': aktif, 'tb': tb, 'ls': ls,
            'keluar': keluar, 'meninggal': meninggal,
        }

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Profil Dosen"

    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    subheader_fill = PatternFill(start_color='2d6a9f', end_color='2d6a9f', fill_type='solid')
    fakultas_fill = PatternFill(start_color='dbeafe', end_color='dbeafe', fill_type='solid')
    total_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Judul
    ws.merge_cells('A1:R1')
    ws['A1'] = 'REKAP PROFIL DOSEN'
    ws['A1'].font = Font(bold=True, size=14, color='1e3a5f')
    ws['A1'].alignment = center

    ws.merge_cells('A2:R2')
    ws['A2'] = 'Universitas Ichsan Gorontalo'
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = center

    # Header Row 4 & 5
    ws.merge_cells('A4:A5')
    ws['A4'] = 'No'
    ws['A4'].font = header_font
    ws['A4'].fill = header_fill
    ws['A4'].alignment = center
    ws['A4'].border = thin

    ws.merge_cells('B4:B5')
    ws['B4'] = 'Fakultas / Program Studi'
    ws['B4'].font = header_font
    ws['B4'].fill = header_fill
    ws['B4'].alignment = center
    ws['B4'].border = thin

    ws.merge_cells('C4:C5')
    ws['C4'] = 'Total Dosen'
    ws['C4'].font = header_font
    ws['C4'].fill = header_fill
    ws['C4'].alignment = center
    ws['C4'].border = thin

    ws.merge_cells('D4:F4')
    ws['D4'] = 'Pendidikan Terakhir'
    ws['D4'].font = header_font
    ws['D4'].fill = header_fill
    ws['D4'].alignment = center
    ws['D4'].border = thin

    ws.merge_cells('G4:L4')
    ws['G4'] = 'Jabatan Fungsional'
    ws['G4'].font = header_font
    ws['G4'].fill = header_fill
    ws['G4'].alignment = center
    ws['G4'].border = thin

    ws.merge_cells('M4:M5')
    ws['M4'] = 'Serdos'
    ws['M4'].font = header_font
    ws['M4'].fill = header_fill
    ws['M4'].alignment = center
    ws['M4'].border = thin

    ws.merge_cells('N4:R4')
    ws['N4'] = 'Status Kepegawaian'
    ws['N4'].font = header_font
    ws['N4'].fill = header_fill
    ws['N4'].alignment = center
    ws['N4'].border = thin

    # Sub header row 5
    pend_sub = ['S1', 'S2', 'S3']
    jabfung_sub = ['Tenaga Pengajar', 'Asisten Ahli', 'Lektor',
                   'Lektor Kepala', 'Guru Besar', 'Belum']
    status_sub = ['Aktif', 'Tugas Belajar', 'Lanjut Studi', 'Keluar', 'Meninggal']

    for col, label in enumerate(pend_sub, start=4):  # D=4
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin

    for col, label in enumerate(jabfung_sub, start=7):  # G=7
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin

    for col, label in enumerate(status_sub, start=14):  # N=14
        cell = ws.cell(row=5, column=col, value=label)
        cell.font = Font(bold=True, color='FFFFFF', size=9)
        cell.fill = subheader_fill
        cell.alignment = center
        cell.border = thin

    ws.row_dimensions[4].height = 25
    ws.row_dimensions[5].height = 35

    # Data
    row_num = 6
    grand = {k: 0 for k in [
        'total', 's1', 's2', 's3',
        'tp', 'aa', 'lektor', 'lk', 'gb',
        'serdos', 'aktif', 'tb', 'ls', 'keluar', 'meninggal'
    ]}
    no = 1

    for fak in fakultas_list:
        fak_counts = get_profil_counts({'kode_fakultas': fak.kode_fakultas})
        belum_jabfung = fak_counts['total'] - (
            fak_counts['tp'] + fak_counts['aa'] +
            fak_counts['lektor'] + fak_counts['lk'] + fak_counts['gb']
        )

        fak_font = Font(bold=True, size=10, color='1e3a5f')
        row_data = [
            no, f'[{fak.kode_fakultas}] {fak.nama_fakultas}',
            fak_counts['total'],
            fak_counts['s1'], fak_counts['s2'], fak_counts['s3'],
            fak_counts['tp'], fak_counts['aa'],
            fak_counts['lektor'], fak_counts['lk'],
            fak_counts['gb'], belum_jabfung,
            fak_counts['serdos'],
            fak_counts['aktif'], fak_counts['tb'],
            fak_counts['ls'], fak_counts['keluar'],
            fak_counts['meninggal'],
        ]
        for col, val in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = fak_font
            cell.fill = fakultas_fill
            cell.alignment = center if col != 2 else left
            cell.border = thin
        ws.row_dimensions[row_num].height = 20
        row_num += 1
        no += 1

        for k in grand:
            grand[k] += fak_counts.get(k, 0)

        # Prodi
        prodi_fak = prodi_list.filter(fakultas=fak)
        for prodi in prodi_fak:
            pc = get_profil_counts({'kode_prodi': prodi.kode_prodi})
            belum_jf = pc['total'] - (
                pc['tp'] + pc['aa'] + pc['lektor'] +
                pc['lk'] + pc['gb']
            )
            row_data_p = [
                '', f'    {prodi.kode_prodi} - {prodi.nama_prodi}',
                pc['total'],
                pc['s1'], pc['s2'], pc['s3'],
                pc['tp'], pc['aa'], pc['lektor'],
                pc['lk'], pc['gb'], belum_jf,
                pc['serdos'],
                pc['aktif'], pc['tb'],
                pc['ls'], pc['keluar'], pc['meninggal'],
            ]
            for col, val in enumerate(row_data_p, start=1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.font = Font(size=9)
                cell.alignment = center if col != 2 else left
                cell.border = thin
            ws.row_dimensions[row_num].height = 18
            row_num += 1

    # Grand Total
    total_font = Font(bold=True, color='FFFFFF', size=10)
    ws.cell(row=row_num, column=1, value='').border = thin
    belum_total = grand['total'] - (
        grand['tp'] + grand['aa'] + grand['lektor'] +
        grand['lk'] + grand['gb']
    )
    total_label = ws.cell(row=row_num, column=2, value='TOTAL KESELURUHAN')
    total_label.font = total_font
    total_label.fill = total_fill
    total_label.alignment = left
    total_label.border = thin

    total_values = [
        grand['total'],
        grand['s1'], grand['s2'], grand['s3'],
        grand['tp'], grand['aa'], grand['lektor'],
        grand['lk'], grand['gb'], belum_total,
        grand['serdos'],
        grand['aktif'], grand['tb'],
        grand['ls'], grand['keluar'], grand['meninggal'],
    ]
    for col, val in enumerate(total_values, start=3):
        cell = ws.cell(row=row_num, column=col, value=val)
        cell.font = total_font
        cell.fill = total_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[row_num].height = 22

    # Lebar kolom
    col_widths = [5, 35, 12, 8, 8, 8, 15, 14, 10, 14, 12,
                  10, 10, 10, 14, 14, 10, 12]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = 'C6'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Rekap_Profil_Dosen.xlsx"'
    wb.save(response)
    return response