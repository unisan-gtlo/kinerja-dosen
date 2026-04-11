from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.hashers import make_password
from master.models import Fakultas, Prodi
from .models import User
import openpyxl
from django.http import HttpResponse
from django.contrib.auth.hashers import make_password
from django.core.paginator import Paginator
from django.http import JsonResponse
import random
import logging
from django.utils import timezone

logger = logging.getLogger(__name__)
    
  

def generate_captcha(request):
    """Generate soal matematika sederhana dan simpan jawaban di session"""
    a = random.randint(1, 10)
    b = random.randint(1, 10)
    ops = ['+', '-', 'x']
    op = random.choice(ops)
    if op == '+':
        answer = a + b
        soal = f"{a} + {b}"
    elif op == '-':
        if a < b:
            a, b = b, a
        answer = a - b
        soal = f"{a} - {b}"
    else:
        answer = a * b
        soal = f"{a} x {b}"
    request.session['captcha_answer'] = answer
    request.session['captcha_soal'] = soal
    return soal


def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard:index')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        captcha_input = request.POST.get('captcha', '').strip()

        # Validasi input kosong
        if not username or not password:
            messages.error(request, 'Username dan password wajib diisi.')
            soal = generate_captcha(request)
            return render(request, 'accounts/login.html', {'captcha_soal': soal})

        # Validasi CAPTCHA
        captcha_answer = request.session.get('captcha_answer')
        try:
            if int(captcha_input) != captcha_answer:
                messages.error(request, 'Jawaban verifikasi salah. Coba lagi.')
                soal = generate_captcha(request)
                return render(request, 'accounts/login.html', {'captcha_soal': soal})
        except (ValueError, TypeError):
            messages.error(request, 'Jawaban verifikasi harus berupa angka.')
            soal = generate_captcha(request)
            return render(request, 'accounts/login.html', {'captcha_soal': soal})

        # Autentikasi
        user = authenticate(request, username=username, password=password)
        if user is not None:
            if user.status_akun == 'aktif':
                login(request, user)
                logger.info(
                    f'LOGIN SUCCESS: {username} dari IP '
                    f'{request.META.get("REMOTE_ADDR")} '
                    f'pada {timezone.now()}'
                )
                next_url = request.GET.get('next', '/')
                return redirect(next_url)
            else:
                messages.error(
                    request,
                    'Akun Anda tidak aktif. Hubungi administrator.'
                )
                logger.warning(
                    f'LOGIN BLOCKED (nonaktif): {username} '
                    f'dari IP {request.META.get("REMOTE_ADDR")}'
                )
        else:
            messages.error(request, 'Username atau password salah.')
            logger.warning(
                f'LOGIN FAILED: {username} '
                f'dari IP {request.META.get("REMOTE_ADDR")}'
            )

        # Generate captcha baru setelah gagal
        soal = generate_captcha(request)
        return render(request, 'accounts/login.html', {'captcha_soal': soal})

    # GET request — generate captcha baru
    soal = generate_captcha(request)
    return render(request, 'accounts/login.html', {'captcha_soal': soal})

def logout_view(request):
    logout(request)
    return redirect('accounts:login')

@login_required
def kelola_user(request):
    if request.user.role != 'admin':
        messages.error(request, 'Anda tidak memiliki akses.')
        return redirect('dashboard:index')

    fakultas_list = Fakultas.objects.filter(status='aktif').order_by('kode_fakultas')
    prodi_list = Prodi.objects.filter(status='aktif').order_by('kode_prodi')

    # Filter
    role_filter = request.GET.get('role', '')
    fakultas_filter = request.GET.get('fakultas', '')
    prodi_filter = request.GET.get('prodi', '')
    cari = request.GET.get('cari', '')

    users = User.objects.all().order_by('kode_fakultas', 'kode_prodi', 'first_name')

    if role_filter:
        users = users.filter(role=role_filter)
    if fakultas_filter:
        users = users.filter(kode_fakultas=fakultas_filter)
    if prodi_filter:
        users = users.filter(kode_prodi=prodi_filter)
    if cari:
        users = users.filter(
            first_name__icontains=cari
        ) | users.filter(
            last_name__icontains=cari
        ) | users.filter(
            username__icontains=cari
        ) | users.filter(
            nidn__icontains=cari
        )

    context = {
        'users': users,
        'fakultas_list': fakultas_list,
        'prodi_list': prodi_list,
        'role_filter': role_filter,
        'fakultas_filter': fakultas_filter,
        'prodi_filter': prodi_filter,
        'cari': cari,
    }
    # Pagination
    per_halaman = int(request.GET.get('per_halaman', 15))
    paginator = Paginator(users, per_halaman)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    context = {
        'users': page_obj,
        'page_obj': page_obj,
        'fakultas_list': fakultas_list,
        'prodi_list': prodi_list,
        'role_filter': role_filter,
        'fakultas_filter': fakultas_filter,
        'prodi_filter': prodi_filter,
        'cari': cari,
        'total': paginator.count,
        'per_halaman': str(per_halaman),
    }
    return render(request, 'accounts/kelola_user.html', context)
   

@login_required
def tambah_user(request):
    if request.user.role != 'admin':
        return redirect('dashboard:index')

    fakultas_list = Fakultas.objects.filter(status='aktif').order_by('kode_fakultas')
    prodi_list = Prodi.objects.filter(status='aktif').order_by('kode_prodi')

    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        first_name = request.POST.get('first_name', '').strip()
        last_name = request.POST.get('last_name', '').strip()
        email = request.POST.get('email', '').strip()
        nidn = request.POST.get('nidn', '').strip()
        role = request.POST.get('role', 'dosen')
        kode_fakultas = request.POST.get('kode_fakultas', '').strip()
        kode_prodi = request.POST.get('kode_prodi', '').strip()
        no_hp = request.POST.get('no_hp', '').strip()
        status_kepegawaian = request.POST.get('status_kepegawaian', 'Aktif')
        password = request.POST.get('password', '').strip()

        if not username or not password or not first_name:
            messages.error(request, 'Username, nama depan, dan password wajib diisi.')
        elif User.objects.filter(username=username).exists():
            messages.error(request, f'Username "{username}" sudah digunakan.')
        else:
            user = User.objects.create(
                username=username,
                first_name=first_name,
                last_name=last_name,
                email=email,
                nidn=nidn,
                role=role,
                kode_fakultas=kode_fakultas,
                kode_prodi=kode_prodi,
                no_hp=no_hp,
                status_kepegawaian=status_kepegawaian,
                password=make_password(password),
                status_akun='aktif'
            )
            messages.success(request, f'User "{username}" berhasil ditambahkan.')
            return redirect('accounts:kelola_user')

    context = {
        'fakultas_list': fakultas_list,
        'prodi_list': prodi_list,
    }
    return render(request, 'accounts/tambah_user.html', context)

@login_required
def edit_user(request, user_id):
    if request.user.role != 'admin':
        return redirect('dashboard:index')

    target_user = get_object_or_404(User, id=user_id)
    fakultas_list = Fakultas.objects.filter(status='aktif').order_by('kode_fakultas')
    prodi_list = Prodi.objects.filter(status='aktif').order_by('kode_prodi')

    if request.method == 'POST':
        target_user.first_name = request.POST.get('first_name', '').strip()
        target_user.last_name = request.POST.get('last_name', '').strip()
        target_user.email = request.POST.get('email', '').strip()
        target_user.nidn = request.POST.get('nidn', '').strip()
        target_user.role = request.POST.get('role', target_user.role)
        target_user.kode_fakultas = request.POST.get('kode_fakultas', '').strip()
        target_user.kode_prodi = request.POST.get('kode_prodi', '').strip()
        target_user.no_hp = request.POST.get('no_hp', '').strip()
        target_user.status_akun = request.POST.get('status_akun', 'aktif')

        password_baru = request.POST.get('password_baru', '').strip()
        if password_baru:
            target_user.password = make_password(password_baru)

        target_user.save()
        messages.success(request, f'User "{target_user.username}" berhasil diupdate.')
        return redirect('accounts:kelola_user')

    context = {
        'target_user': target_user,
        'fakultas_list': fakultas_list,
        'prodi_list': prodi_list,
    }
    return render(request, 'accounts/edit_user.html', context)

@login_required
def hapus_user(request, user_id):
    if request.user.role != 'admin':
        return redirect('dashboard:index')

    target_user = get_object_or_404(User, id=user_id)
    if target_user == request.user:
        messages.error(request, 'Tidak bisa menghapus akun sendiri.')
    else:
        username = target_user.username
        target_user.delete()
        messages.success(request, f'User "{username}" berhasil dihapus.')
    return redirect('accounts:kelola_user')



@login_required
def import_user(request):
    if request.user.role != 'admin':
        return redirect('dashboard:index')

    if request.method == 'POST':
        if 'file_excel' not in request.FILES:
            messages.error(request, 'Pilih file Excel terlebih dahulu.')
            return redirect('accounts:kelola_user')

        file = request.FILES['file_excel']
        if not file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'File harus berformat Excel (.xlsx)')
            return redirect('accounts:kelola_user')

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active

            berhasil = 0
            gagal = 0
            error_list = []

            for row_num, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
                # Skip baris kosong
                if not row[0] and not row[1]:
                    continue

                try:
                    username = str(row[0]).strip() if row[0] else ''
                    nidn = str(row[1]).strip() if row[1] else ''
                    first_name = str(row[2]).strip() if row[2] else ''
                    last_name = str(row[3]).strip() if row[3] else ''
                    email = str(row[4]).strip() if row[4] else ''
                    no_hp = str(row[5]).strip() if row[5] else ''
                    role = str(row[6]).strip().lower() if row[6] else 'dosen'
                    kode_fakultas = str(row[7]).strip().upper() if row[7] else ''
                    kode_prodi = str(row[8]).strip().upper() if row[8] else ''
                    status_kepegawaian = str(row[9]).strip() if row[9] else 'Aktif'
                    password = str(row[10]).strip() if row[10] else 'unichsan123'

                    if not username:
                        error_list.append(f'Baris {row_num}: Username kosong')
                        gagal += 1
                        continue

                    if User.objects.filter(username=username).exists():
                        error_list.append(f'Baris {row_num}: Username "{username}" sudah ada')
                        gagal += 1
                        continue

                    # Validasi role
                    valid_roles = ['admin', 'rektorat', 'biro', 'dekan', 'wadek',
                                   'kaprodi', 'sekprodi', 'operator', 'dosen']
                    if role not in valid_roles:
                        role = 'dosen'

                    User.objects.create(
                        username=username,
                        nidn=nidn,
                        first_name=first_name,
                        last_name=last_name,
                        email=email,
                        no_hp=no_hp,
                        role=role,
                        kode_fakultas=kode_fakultas,
                        kode_prodi=kode_prodi,
                        status_kepegawaian=status_kepegawaian,
                        status_akun='aktif',
                        password=make_password(password),
                    )
                    berhasil += 1

                except Exception as e:
                    error_list.append(f'Baris {row_num}: {str(e)}')
                    gagal += 1

            if berhasil > 0:
                messages.success(request, f'{berhasil} data berhasil diimport.')
            if gagal > 0:
                messages.warning(request, f'{gagal} data gagal: {", ".join(error_list[:5])}')

        except Exception as e:
            messages.error(request, f'Gagal membaca file: {str(e)}')

    return redirect('accounts:kelola_user')


@login_required
def download_template(request):
    if request.user.role != 'admin':
        return redirect('dashboard:index')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Import Dosen"

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    info_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    info_font = Font(italic=True, size=10, color='555555')

    # Judul
    ws.merge_cells('A1:K1')
    ws['A1'] = 'TEMPLATE IMPORT DATA PENGGUNA - SIKD UNIVERSITAS ICHSAN GORONTALO'
    ws['A1'].font = Font(bold=True, size=13, color='1e3a5f')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 25

    # Header kolom
    headers = [
        'Username *', 'NIDN', 'Nama Depan *', 'Nama Belakang',
        'Email', 'No. HP', 'Role *', 'Kode Fakultas',
        'Kode Prodi', 'Status Kepegawaian', 'Password *'
    ]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[2].height = 25

    # Contoh data
    contoh = [
        ['dosen001', '0123456789', 'Ahmad', 'Fauzi', 'ahmad@unichsan.ac.id',
         '081234567890', 'dosen', 'FK', 'T31', 'Aktif', 'unichsan123'],
        ['dosen002', '9876543210', 'Siti', 'Rahayu', 'siti@unichsan.ac.id',
         '082345678901', 'dosen', 'FE', 'E21', 'Aktif', 'unichsan123'],
        ['kaprodi001', '1122334455', 'Budi', 'Santoso', 'budi@unichsan.ac.id',
         '083456789012', 'kaprodi', 'FH', 'H11', 'Aktif', 'unichsan123'],
    ]

    example_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
    for row_idx, data in enumerate(contoh, 3):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill = example_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[row_idx].height = 20

    # Baris kosong untuk input data (10 baris)
    for row_idx in range(6, 16):
        for col_idx in range(1, 12):
            cell = ws.cell(row=row_idx, column=col_idx, value='')
            cell.border = thin
            cell.alignment = center
        ws.row_dimensions[row_idx].height = 20

    # Keterangan
    keterangan_row = 17
    ws.merge_cells(f'A{keterangan_row}:K{keterangan_row}')
    ws[f'A{keterangan_row}'] = 'KETERANGAN:'
    ws[f'A{keterangan_row}'].font = Font(bold=True, size=10)

    keterangans = [
        ('A', f'A{keterangan_row+1}:K{keterangan_row+1}',
         '* = Wajib diisi. Baris hijau adalah contoh data — hapus sebelum import.'),
        ('B', f'A{keterangan_row+2}:K{keterangan_row+2}',
         'Role yang valid: admin, rektorat, biro, dekan, wadek, kaprodi, sekprodi, operator, dosen'),
        ('C', f'A{keterangan_row+3}:K{keterangan_row+3}',
         'Kode Fakultas: FE, FH, FISIP, FK, FP, FT, S2'),
        ('D', f'A{keterangan_row+4}:K{keterangan_row+4}',
         'Status Kepegawaian: Aktif, Tugas Belajar, Lanjut Studi, Keluar, Meninggal'),
        ('E', f'A{keterangan_row+5}:K{keterangan_row+5}',
         'Password default jika dikosongkan: unichsan123 (dapat diubah setelah login)'),
        ('F', f'A{keterangan_row+6}:K{keterangan_row+6}',
         'Username tidak boleh mengandung spasi dan harus unik (belum terdaftar)'),
    ]

    for _, merge_range, text in keterangans:
        ws.merge_cells(merge_range)
        start_cell = merge_range.split(':')[0]
        ws[start_cell] = f'• {text}'
        ws[start_cell].font = info_font
        ws[start_cell].fill = info_fill
        ws[start_cell].alignment = Alignment(horizontal='left', vertical='center')

    # Lebar kolom
    col_widths = [15, 15, 18, 18, 25, 15, 12, 15, 12, 20, 15]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Response download
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Template_Import_Pengguna_SIKD.xlsx"'
    wb.save(response)
    return response

@login_required
def ganti_password(request):
    if request.method == 'POST':
        password_lama = request.POST.get('password_lama', '')
        password_baru = request.POST.get('password_baru', '')
        password_konfirmasi = request.POST.get('password_konfirmasi', '')

        user = request.user

        if not user.check_password(password_lama):
            messages.error(request, 'Password lama tidak sesuai.')
        elif len(password_baru) < 8:
            messages.error(request, 'Password baru minimal 8 karakter.')
        elif password_baru == password_lama:
            messages.error(request, 'Password baru tidak boleh sama dengan password lama.')
        elif password_baru != password_konfirmasi:
            messages.error(request, 'Konfirmasi password tidak cocok.')
        elif password_baru.isdigit():
            messages.error(request, 'Password tidak boleh hanya angka.')
        elif password_baru.isalpha():
            messages.error(request, 'Password harus mengandung minimal 1 angka.')
        else:
            user.set_password(password_baru)
            user.save()
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)
            logger.info(
                f'PASSWORD CHANGED: {user.username} '
                f'pada {timezone.now()}'
            )
            messages.success(request, 'Password berhasil diubah.')
            return redirect('accounts:ganti_password')

    return render(request, 'accounts/ganti_password.html')

@login_required
def unlock_user(request, user_id):
    if request.user.role != 'admin':
        return redirect('dashboard:index')
    
    target_user = get_object_or_404(User, id=user_id)
    
    try:
        from axes.models import AccessAttempt
        AccessAttempt.objects.filter(
            username=target_user.username
        ).delete()
        messages.success(
            request,
            f'Akun "{target_user.username}" berhasil dibuka kuncinya.'
        )
    except Exception as e:
        messages.error(request, f'Gagal membuka kunci: {str(e)}')
    
    return redirect('accounts:kelola_user')



def refresh_captcha(request):
    soal = generate_captcha(request)
    return JsonResponse({'soal': soal})
